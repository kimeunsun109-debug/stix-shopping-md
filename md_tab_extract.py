# -*- coding: utf-8 -*-
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

OUT = Path(__file__).parent
TODAY = datetime.now().strftime("%Y-%m-%d")
SKIP = (
    "로그인", "로그아웃", "전체", "선택", "매뉴얼", "FAQ", "Open API", "광고",
    "예약 관리", "여행", "숙박", "렌터카", "Copyright", "약관", "고객센터",
    "Top으로", "사업자", "검색어", "판매상태", "번호", "조회", "등록",
)


def clean(s):
    return re.sub(r"\s+", " ", str(s or "").strip())


def is_product_line(t: str) -> bool:
    if len(t) < 12 or len(t) > 200:
        return False
    if any(x in t for x in SKIP):
        return False
    if re.fullmatch(r"[\d,원\s~.-]+", t):
        return False
    # 상품명 후보: 한글 포함 + 키워드 또는 충분한 길이
    if re.search(r"[가-힣]", t):
        keys = ("보석", "십자", "스팃", "DIY", "접착", "본드", "패키지", "공예", "키트", "E6000", "B6000", "액자", "자수", "마크라", "스티커")
        return any(k in t for k in keys)
    return False


def js_collect():
    return """() => {
        const o = [];
        document.querySelectorAll('td, a, span, div, input, p, li, label, strong').forEach(el => {
            let t = (el.value || el.innerText || el.title || '').trim().replace(/\\s+/g, ' ');
            if (t.length > 12 && t.length < 200) o.push(t);
        });
        return [...new Set(o)];
    }"""


def extract_esm(page):
    page.wait_for_timeout(2000)
    if "goods-manage" not in page.url:
        page.goto("https://www.esmplus.com/Home/v2/goods-manage", timeout=60000)
        page.wait_for_timeout(5000)
    # 검색어 필드 (상품명/브랜드)
    for sel in [
        'input[placeholder*="상품"]',
        'input[placeholder*="검색"]',
        'textarea',
        'input[type="text"]:visible',
    ]:
        try:
            loc = page.locator(sel)
            for i in range(min(loc.count(), 5)):
                el = loc.nth(i)
                if el.is_visible():
                    el.fill("보석십자수")
                    page.wait_for_timeout(500)
                    break
        except Exception:
            pass
    for sel in ['button:has-text("검색")', 'button:has-text("조회")', 'a:has-text("검색")']:
        try:
            btn = page.locator(sel).last
            if btn.count() and btn.is_visible():
                btn.click(timeout=5000)
                page.wait_for_timeout(10000)
                break
        except Exception:
            pass
    return collect_all(page)


def extract_11st(page):
    page.wait_for_timeout(2000)
    for sel in ['button:has-text("검색")', 'button:has-text("조회")', 'input[type="button"][value*="검색"]']:
        try:
            btn = page.locator(sel).first
            if btn.count() and btn.is_visible():
                btn.click(timeout=3000)
                page.wait_for_timeout(8000)
                break
        except Exception:
            pass
    page.evaluate("window.scrollTo(0, 800)")
    page.wait_for_timeout(2000)
    return collect_all(page)


def collect_all(page):
    found = []
    for fr in page.frames:
        try:
            for t in fr.evaluate(js_collect()):
                t = clean(t)
                if is_product_line(t) and t not in found:
                    found.append(t)
        except Exception:
            pass
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    page.wait_for_timeout(1500)
    for fr in page.frames:
        try:
            for t in fr.evaluate(js_collect()):
                t = clean(t)
                if is_product_line(t) and t not in found:
                    found.append(t)
        except Exception:
            pass
    return found[:50]


results = {}
with sync_playwright() as p:
    browser = p.chromium.connect_over_cdp("http://127.0.0.1:9233")
    for pg in browser.contexts[0].pages:
        u = pg.url
        print("tab:", u[:90])
        if "esmplus.com" in u and ("goods-manage" in u or "order-integration" in u or "manage" in u):
            results["지마켓/옥션"] = extract_esm(pg)
            try:
                pg.screenshot(path=str(OUT / "screenshots" / "esm_goods.png"), timeout=10000)
            except Exception:
                pass
        elif "soffice.11st" in u:
            results["11번가"] = extract_11st(pg)
            try:
                pg.screenshot(path=str(OUT / "screenshots" / "11st_goods.png"), timeout=10000)
            except Exception:
                pass

for k, v in results.items():
    print(k, len(v))
    for x in v[:8]:
        print(" ", x[:95])

if results:
    xlsx = OUT / f"STIX_MD_실데이터_{TODAY}.xlsx"
    wb = load_workbook(xlsx) if xlsx.exists() else load_workbook()
    if "지옥션_11번가_추가" in wb.sheetnames:
        del wb["지옥션_11번가_추가"]
    ws = wb.create_sheet("지옥션_11번가_추가")
    ws.append(["플랫폼", "상품명"])
    for plat, prods in results.items():
        for pr in prods:
            ws.append([plat, pr])
    wb.save(xlsx)
    print("saved", xlsx)
