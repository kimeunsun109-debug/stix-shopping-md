# -*- coding: utf-8 -*-
"""로켓그로스 6월 실마진 — 정산액 - 원가 - 보관비 - 입출고비 - 배송(반출)비"""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent
RG_DIR = BASE / "쇼핑몰별 전체상품" / "로켓그로스"
TODAY = datetime.now().strftime("%Y-%m-%d")
OUT_XLSX = BASE / f"STIX_로켓그로스_실마진_{TODAY}.xlsx"
OUT_TXT = BASE / f"MD_로켓그로스_실마진_{TODAY}.txt"

# 사용자 제공 원가 (원)
COGS = {
    "15ml_x3": 1080,
    "15ml_x5": 1800,
    "110ml_x1": 2600,
    "110ml_x2": 5200,
    "110ml_x4": 5200,
    "cross_stitch_pkg": 7130,
    "bead_sticker": 1800,
    "bead_frame": 2900,
    "bead_kit": 5500,  # 패브릭형·재물운 등 보석십자수 DIY
}


def clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def num(v, default=0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def find_header_row(rows: list[tuple], marker: str = "과금유형") -> int | None:
    markers = {marker, "과금유형", "정산유형"}
    for i, r in enumerate(rows[:20]):
        if r and clean(r[0]) in markers:
            return i
    return None


def merge_headers(h1: list, h2: list | None) -> list[str]:
    out = []
    last = ""
    for i, a in enumerate(h1):
        a = clean(a)
        b = clean(h2[i]) if h2 and i < len(h2) else ""
        if a and b:
            out.append(f"{a}|{b}")
        elif a:
            out.append(a)
            last = a
        elif b:
            out.append(f"{last}|{b}" if last else b)
        else:
            out.append("")
    return out


def col_idx(headers: list[str], *needles: str) -> int | None:
    for i, h in enumerate(headers):
        if all(n in h for n in needles):
            return i
    return None


def settle_col(headers: list[str]) -> int | None:
    idx = col_idx(headers, "판매수수료금액")
    if idx is not None and idx + 1 < len(headers):
        return idx + 1
    for i, h in enumerate(headers):
        if not h:
            continue
        if h.startswith("정산") and h.endswith("액"):
            if "정산금액" in h or "정산예정금액" in h:
                continue
            return i
    return None


def fee_col(headers: list[str], kind: str = "") -> int | None:
    for needles in (
        ("보관료부과액",),
        ("최종청구금액",),
        ("물류센터부과액",),
        ("A-B-C",),
        ("부과액",),
    ):
        idx = col_idx(headers, *needles)
        if idx is not None:
            return idx
    fallback = {"storage": 19, "ship": 24, "inout": 19}
    fb = fallback.get(kind)
    if fb is not None and len(headers) > fb:
        return fb
    return None


def load_sheet_rows(path: Path, max_scan: int = 25) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(max_row=50000, values_only=True))
    wb.close()
    return rows


def is_subheader_row(row: tuple) -> bool:
    if not row:
        return False
    cells = [clean(c) for c in row if c]
    if not cells:
        return False
    if row[0] and clean(row[0]) in ("판매수수료", "WEEKLY", "주정산", "주문정산", "정산"):
        return False
    if any("(A-B-C)" in c or "(A-B)" in c for c in cells):
        return True
    return sum(1 for c in row[:10] if c) == 0 and any("(A)" in c for c in cells)


def parse_table(path: Path, header_marker: str = "과금유형") -> tuple[list[str], list[tuple]]:
    rows = load_sheet_rows(path)
    hi = find_header_row(rows, header_marker)
    if hi is None:
        return [], []
    h1 = [clean(c) for c in rows[hi]]
    h2 = None
    if hi + 1 < len(rows) and is_subheader_row(rows[hi + 1]):
        h2 = [clean(c) for c in rows[hi + 1]]
    headers = merge_headers(h1, h2)
    data_start = hi + 2 if h2 else hi + 1
    data = [r for r in rows[data_start:] if r and r[0] and str(r[0]) not in ("", "합계")]
    return headers, data


def fee_from_row(headers: list[str], row: tuple, *fee_needles: str) -> float:
    idx = col_idx(headers, *fee_needles)
    if idx is None:
        idx = col_idx(headers, "부과액")
    return num(row[idx]) if idx is not None else 0.0


SALES_ROW_TYPES = {"판매수수료", "주문정산", "주정산", "정산"}


def is_sales_row(row: tuple) -> bool:
    return clean(row[0]) in SALES_ROW_TYPES


def map_cogs(name: str, opt: str) -> tuple[str, int]:
    text = f"{name} {opt}"
    if "십자수" in name or "십자수" in opt:
        if "스티커" in text:
            return "보석십자수 스티커", COGS["bead_sticker"]
        if "액자" in text:
            return "보석십자수 액자형", COGS["bead_frame"]
        if "패키지" in text or "프린트" in name or "도구" in name:
            return "십자수 패키지", COGS["cross_stitch_pkg"]
        return "보석십자수 패브릭/기타", COGS["bead_kit"]
    if "B7000" in text.upper():
        if re.search(r"110\s*ml.*[×xX]\s*4|110\s*ml.*4\s*개", text, re.I):
            return "B7000 110ml×4", COGS["110ml_x4"]
        if re.search(r"110\s*ml.*[×xX]\s*2|110\s*ml.*2\s*개", text, re.I):
            return "B7000 110ml×2", COGS["110ml_x2"]
        if re.search(r"110\s*ml", text, re.I):
            return "B7000 110ml×1", COGS["110ml_x1"]
        if re.search(r"15\s*ml.*[×xX]\s*5|15\s*ml.*5\s*개", text, re.I):
            return "B7000 15ml×5", COGS["15ml_x5"]
        if re.search(r"15\s*ml.*[×xX]\s*3|15\s*ml.*3\s*개", text, re.I):
            return "B7000 15ml×3", COGS["15ml_x3"]
    return "기타(원가미등록)", 0


def sku_key(row: tuple, headers: list[str]) -> str:
    sku_i = col_idx(headers, "SKU ID")
    opt_i = col_idx(headers, "옵션명")
    if sku_i is not None and row[sku_i]:
        return str(row[sku_i])
    if opt_i is not None:
        return clean(row[opt_i])
    return ""


def load_sales() -> dict:
    by_sku: dict = defaultdict(
        lambda: {
            "name": "",
            "opt": "",
            "qty": 0.0,
            "sales": 0.0,
            "settle": 0.0,
            "cogs_key": "",
            "cogs_unit": 0,
        }
    )
    for p in sorted(RG_DIR.glob("*.xlsx")):
        if not p.name.startswith("로켓그로스 6월"):
            continue
        if any(x in p.name for x in ("배송", "보관", "입출")):
            continue
        if "정산" in p.name and "로켓그로스" not in p.name:
            continue
        headers, data = parse_table(p, "정산유형")
        if not headers:
            continue
        qty_i = col_idx(headers, "판매수량")
        sales_i = col_idx(headers, "판매액")
        settle_i = settle_col(headers)
        name_i = col_idx(headers, "등록상품명")
        opt_i = col_idx(headers, "옵션명")
        for row in data:
            if not is_sales_row(row):
                continue
            key = sku_key(row, headers)
            if not key:
                continue
            name = clean(row[name_i]) if name_i is not None else ""
            opt = clean(row[opt_i]) if opt_i is not None else ""
            qty = num(row[qty_i]) if qty_i is not None else 0
            if qty <= 0:
                continue
            rec = by_sku[key]
            rec["name"] = name or rec["name"]
            rec["opt"] = opt or rec["opt"]
            rec["qty"] += qty
            rec["sales"] += num(row[sales_i]) if sales_i is not None else 0
            rec["settle"] += num(row[settle_i]) if settle_i is not None else 0
            ck, cu = map_cogs(rec["name"], rec["opt"])
            rec["cogs_key"] = ck
            rec["cogs_unit"] = cu
    return by_sku


def load_fee_by_sku(kind: str) -> dict[str, float]:
    """kind: ship | storage | inout"""
    fees: dict[str, float] = defaultdict(float)
    patterns = {
        "ship": "*배송*.xlsx",
        "storage": "*보관*.xlsx",
        "inout": "6*.xlsx",
    }
    marker = {"ship": "과금유형", "storage": "과금유형", "inout": "과금유형"}
    for p in sorted(RG_DIR.glob(patterns[kind])):
        if kind != "inout" and not p.name.startswith("로켓그로스"):
            continue
        if kind == "inout" and p.name.startswith("로켓그로스"):
            continue
        headers, data = parse_table(p, marker[kind])
        if not headers:
            continue
        fee_i = fee_col(headers, kind)
        for row in data:
            key = sku_key(row, headers)
            if not key or fee_i is None:
                continue
            fees[key] += num(row[fee_i])
    return fees


def build_report():
    sales = load_sales()
    ship = load_fee_by_sku("ship")
    storage = load_fee_by_sku("storage")
    inout = load_fee_by_sku("inout")

    rows = []
    totals = defaultdict(float)
    for sku, s in sorted(sales.items(), key=lambda x: -x[1]["settle"]):
        qty = s["qty"]
        cogs_total = s["cogs_unit"] * qty
        ship_fee = ship.get(sku, 0)
        stor_fee = storage.get(sku, 0)
        inout_fee = inout.get(sku, 0)
        margin = s["settle"] - cogs_total - ship_fee - stor_fee - inout_fee
        margin_rate = (margin / s["sales"] * 100) if s["sales"] else 0
        per_unit = margin / qty if qty else 0
        rows.append(
            {
                "SKU": sku,
                "상품명": s["name"],
                "옵션": s["opt"],
                "원가구분": s["cogs_key"],
                "판매수량": int(qty),
                "판매액": int(s["sales"]),
                "정산액": int(s["settle"]),
                "원가합계": int(cogs_total),
                "배송반출비": int(ship_fee),
                "보관비": int(stor_fee),
                "입출고비": int(inout_fee),
                "실마진": int(margin),
                "실마진율%": round(margin_rate, 1),
                "건당실마진": int(per_unit),
            }
        )
        for k, v in [
            ("qty", qty),
            ("sales", s["sales"]),
            ("settle", s["settle"]),
            ("cogs", cogs_total),
            ("ship", ship_fee),
            ("storage", stor_fee),
            ("inout", inout_fee),
            ("margin", margin),
        ]:
            totals[k] += v

    return rows, totals


def write_outputs(rows: list[dict], totals: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "실마진"
    headers = list(rows[0].keys()) if rows else []
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for c in ws[1]:
        c.fill = hf
        c.font = hfont
    for r in rows:
        ws.append([r[h] for h in headers])
    ws.append([])
    ws.append(
        [
            "합계",
            "",
            "",
            "",
            int(totals["qty"]),
            int(totals["sales"]),
            int(totals["settle"]),
            int(totals["cogs"]),
            int(totals["ship"]),
            int(totals["storage"]),
            int(totals["inout"]),
            int(totals["margin"]),
            round(totals["margin"] / totals["sales"] * 100, 1) if totals["sales"] else 0,
            "",
        ]
    )
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 40
    wb.save(OUT_XLSX)

    lines = [
        f"로켓그로스 6월 실마진 분석 ({TODAY})",
        "실마진 = 정산액 - 원가 - 배송(반출)비 - 보관비 - 입출고비",
        "",
        "[적용 원가]",
        f"  B7000 15ml×3: {COGS['15ml_x3']:,}원",
        f"  B7000 15ml×5: {COGS['15ml_x5']:,}원",
        f"  B7000 110ml×1: {COGS['110ml_x1']:,}원",
        f"  B7000 110ml×2: {COGS['110ml_x2']:,}원",
        f"  B7000 110ml×4: {COGS['110ml_x4']:,}원",
        f"  십자수 패키지: {COGS['cross_stitch_pkg']:,}원",
        f"  보석십자수 스티커: {COGS['bead_sticker']:,}원",
        f"  보석십자수 액자형: {COGS['bead_frame']:,}원",
        f"  보석십자수 패브릭/기타: {COGS['bead_kit']:,}원",
        "",
        "[6월 전체 합계]",
        f"  판매수량: {int(totals['qty']):,}건",
        f"  판매액: {int(totals['sales']):,}원",
        f"  정산액: {int(totals['settle']):,}원",
        f"  원가: {int(totals['cogs']):,}원",
        f"  배송(반출)비: {int(totals['ship']):,}원",
        f"  보관비: {int(totals['storage']):,}원",
        f"  입출고비: {int(totals['inout']):,}원",
        f"  ★ 실마진: {int(totals['margin']):,}원 ({totals['margin']/totals['sales']*100:.1f}%)" if totals["sales"] else "",
        "",
        "[SKU별 TOP (실마진 기준)]",
    ]
    for r in sorted(rows, key=lambda x: -x["실마진"])[:25]:
        lines.append(
            f"  {r['원가구분']:<16} qty={r['판매수량']:>4}  "
            f"정산 {r['정산액']:>9,}  실마진 {r['실마진']:>8,} ({r['실마진율%']}%)  "
            f"건당 {r['건당실마진']:>6,}원"
        )
    unmapped = [r for r in rows if r["원가구분"] == "기타(원가미등록)"]
    if unmapped:
        lines.extend(["", "[원가 미등록 SKU — 확인 필요]", ""])
        for r in unmapped[:15]:
            lines.append(f"  {r['상품명'][:40]} | {r['옵션'][:30]} | qty={r['판매수량']}")
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")


def main():
    rows, totals = build_report()
    if not rows:
        print("No sales data found")
        return
    write_outputs(rows, totals)
    print(f"Saved: {OUT_XLSX}")
    print(f"Saved: {OUT_TXT}")
    print(
        f"Total margin: {int(totals['margin']):,} / sales {int(totals['sales']):,} "
        f"({totals['margin']/totals['sales']*100:.1f}%)"
    )


if __name__ == "__main__":
    main()
