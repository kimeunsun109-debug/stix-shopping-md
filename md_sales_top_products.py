# -*- coding: utf-8 -*-
"""사이트별 매출 파일에서 판매상위상품 추출"""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent
SALES_DIR = BASE / "쇼핑몰별 전체상품"
TODAY = datetime.now().strftime("%Y-%m-%d")
OUT_XLSX = BASE / f"STIX_사이트별_판매상위_{TODAY}.xlsx"
OUT_TXT = BASE / f"MD_사이트별_판매상위_{TODAY}.txt"

TOP_N = 30
SKIP_NAMES = re.compile(r"배송비|배송료|기본배송|추가배송|도서산간|무료배송|<", re.I)


def clean(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    t = re.sub(r"\s+", " ", str(s).strip())
    return "" if t.lower() == "nan" else t


def num(v, default=0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def skip_product(name: str) -> bool:
    n = clean(name)
    if not n or len(n) < 2:
        return True
    return bool(SKIP_NAMES.search(n))


def add_agg(agg: dict, name: str, qty: float, revenue: float):
    name = clean(name)
    if skip_product(name):
        return
    rec = agg[name]
    rec["qty"] += qty
    rec["revenue"] += revenue
    rec["orders"] += 1 if qty else 0


def norm_col_name(c) -> str:
    return re.sub(r"\s+", " ", str(c).replace("\u00a0", " ")).strip()


def read_smartstore(files: list[Path]) -> dict:
    agg: dict = defaultdict(lambda: {"qty": 0.0, "revenue": 0.0, "orders": 0})
    for p in files:
        df = pd.read_excel(p)
        cols = list(df.columns)
        name_col = cols[4] if len(cols) > 4 else None
        qty_col = cols[5] if len(cols) > 5 else None
        amt_col = cols[6] if len(cols) > 6 else None
        if not name_col or not amt_col:
            continue
        for _, row in df.iterrows():
            name = clean(row.get(name_col, ""))
            revenue = num(row.get(amt_col, 0))
            qty = num(row.get(qty_col, 1)) if qty_col else 1.0
            if revenue <= 0 and qty <= 0:
                continue
            add_agg(agg, name, qty, revenue)
    return agg


def read_esm(files: list[Path]) -> dict:
    agg: dict = defaultdict(lambda: {"qty": 0.0, "revenue": 0.0, "orders": 0})
    for p in files:
        df = pd.read_excel(p)
        for _, row in df.iterrows():
            name = clean(row.get("상품명", ""))
            qty = num(row.get("수량", 0))
            revenue = num(row.get("판매금액", 0))
            if qty <= 0 and revenue <= 0:
                continue
            add_agg(agg, name, qty, revenue)
    return agg


def read_coupang(files: list[Path]) -> dict:
    agg: dict = defaultdict(lambda: {"qty": 0.0, "revenue": 0.0, "orders": 0})
    for p in files:
        df = pd.read_excel(p)
        for _, row in df.iterrows():
            opt = clean(row.get("옵션 ID", ""))
            if "배송" in opt:
                continue
            name = clean(row.get("상품명", ""))
            if skip_product(name):
                continue
            qty = num(row.get("판매수량", 0)) - num(row.get("환불수량", 0))
            revenue = num(row.get("판매액", 0))
            if qty <= 0 and revenue <= 0:
                continue
            add_agg(agg, name, max(qty, 0), revenue)
    return agg


def read_cafe24(files: list[Path]) -> dict:
    agg: dict = defaultdict(lambda: {"qty": 0.0, "revenue": 0.0, "orders": 0})
    for p in files:
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                df = pd.read_csv(p, encoding=enc)
                break
            except UnicodeDecodeError:
                df = None
        if df is None:
            continue
        name_col = "주문상품명" if "주문상품명" in df.columns else "상품명"
        for _, row in df.iterrows():
            name = clean(row.get(name_col, ""))
            qty = num(row.get("수량", 0))
            price = num(row.get("판매가", 0))
            revenue = price * qty if price and qty else price
            if qty <= 0 and revenue <= 0:
                continue
            add_agg(agg, name, qty, revenue)
    return agg


def read_11st(files: list[Path]) -> dict:
    agg: dict = defaultdict(lambda: {"qty": 0.0, "revenue": 0.0, "orders": 0})
    for p in files:
        df = pd.read_excel(p, engine="xlrd", header=5)
        name_col = "상품명" if "상품명" in df.columns else None
        if not name_col:
            continue
        qty_col = "수량" if "수량" in df.columns else None
        rev_col = None
        for c in ("판매금액합계", "판매금액", "결제금액"):
            if c in df.columns:
                rev_col = c
                break
        for _, row in df.iterrows():
            name = clean(row.get(name_col, ""))
            qty = num(row.get(qty_col, 0)) if qty_col else 1.0
            revenue = num(row.get(rev_col, 0)) if rev_col else 0.0
            if qty <= 0 and revenue <= 0:
                continue
            add_agg(agg, name, qty, revenue)
    return agg


def top_rows(agg: dict, n: int = TOP_N) -> list[dict]:
    rows = []
    for name, v in agg.items():
        rows.append(
            {
                "상품명": name,
                "판매수량": int(v["qty"]),
                "매출액": int(v["revenue"]),
                "건수": int(v["orders"]),
            }
        )
    rows.sort(key=lambda x: (-x["매출액"], -x["판매수량"]))
    return rows[:n]


def write_xlsx(platform_data: dict[str, list[dict]]):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for platform, rows in platform_data.items():
        ws = wb.create_sheet(platform.replace("/", "-")[:31])
        headers = ["순위", "상품명", "판매수량", "매출액", "건수"]
        ws.append(headers)
        for c in ws[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        for i, r in enumerate(rows, 1):
            ws.append([i, r["상품명"], r["판매수량"], r["매출액"], r["건수"]])
        ws.column_dimensions["B"].width = 60
        for col in ("A", "C", "D", "E"):
            ws.column_dimensions[col].width = 12
    wb.save(OUT_XLSX)


def write_txt(platform_data: dict[str, list[dict]]):
    lines = [f"STIX 사이트별 판매상위 TOP{TOP_N} ({TODAY})", "=" * 60, ""]
    for platform, rows in platform_data.items():
        lines.append(f"## {platform}")
        lines.append(f"{'순위':>4}  {'매출액':>12}  {'수량':>8}  상품명")
        lines.append("-" * 80)
        for i, r in enumerate(rows, 1):
            lines.append(
                f"{i:4d}  {r['매출액']:>12,}  {r['판매수량']:>8,}  {r['상품명'][:55]}"
            )
        lines.append("")
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")


def main():
    platforms = {
        "스마트스토어": read_smartstore(sorted(SALES_DIR.glob("스마트스토어 *매출*.xlsx"))),
        "지마켓/옥션": read_esm(sorted(SALES_DIR.glob("지마켓,옥션 *매출*.xlsx"))),
        "쿠팡": read_coupang(sorted(SALES_DIR.glob("쿠팡*매출*.xlsx"))),
        "11번가": read_11st(sorted(SALES_DIR.glob("11번가 *매출*.xls"))),
        "카페24": read_cafe24(sorted(SALES_DIR.glob("카페24 *매출*.csv"))),
    }
    platform_data = {k: top_rows(v) for k, v in platforms.items()}
    write_xlsx(platform_data)
    write_txt(platform_data)
    print(f"Saved: {OUT_XLSX}")
    print(f"Saved: {OUT_TXT}")
    for k, rows in platform_data.items():
        print(f"{k}: {len(rows)} products (agg {sum(r['매출액'] for r in rows):,}원 top{TOP_N})")


if __name__ == "__main__":
    main()
