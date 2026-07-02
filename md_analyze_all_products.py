# -*- coding: utf-8 -*-
"""STIX MD - 쇼핑몰별 전체상품 엑셀/CSV 분석"""
import csv
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SRC = Path(__file__).parent / "쇼핑몰별 전체상품"
OUT = Path(__file__).parent
TODAY = datetime.now().strftime("%Y-%m-%d")

FILES = {
    "쿠팡": SRC / "쿠팡.xlsx",
    "로켓그로스": SRC / "로켓배송.로켓그로스.xlsx",
    "지마켓/옥션": SRC / "지마켓,옥션.xlsx",
    "스마트스토어": SRC / "스마트스토어.csv",
    "11번가": SRC / "11번가.xlsx",
    "카페24": SRC / "카페24.자사몰.xlsx",
}


def clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def read_xlsx(path: Path) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return [], []
    header_idx = 0
    for i, row in enumerate(all_rows[:10]):
        cells = [clean(c) for c in row if c]
        if any("등록상품명" in c or c == "상품명" for c in cells):
            header_idx = i
            break
    headers = [clean(h) for h in all_rows[header_idx]]
    data = []
    for row in all_rows[header_idx + 1 :]:
        if not any(row):
            continue
        data.append({headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]})
    return headers, data


def read_csv(path: Path) -> tuple[list[str], list[dict]]:
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        data = [dict(r) for r in reader if any(r.values())]
    return list(headers), data


def find_name_col(headers: list[str], row: dict) -> str:
    for key in (
        "등록상품명", "쿠팡 노출상품명", "상품명", "노출상품명", "ProductName", "productName",
        "판매자상품명", "스마트스토어전용 상품명", "상품 이름",
    ):
        for h in headers:
            if key in h and row.get(h):
                return clean(row.get(h))
    for h in headers:
        if "상품명" in h or "상품 명" in h:
            v = clean(row.get(h))
            if v:
                return v
    return ""


def analyze_name(name: str, platform: str) -> tuple[str, str, str]:
    """returns priority, problem, fix"""
    if not name:
        return "B", "상품명 없음", "상품명 입력 필요"
    problems, fixes = [], []
    n = name
    pl = platform.lower()

    if "스마트" in platform or "카페24" in platform:
        if n.count(",") > 2 or n.count(" ") > 12:
            problems.append("키워드 나열형")
            fixes.append("자연문장형으로 (스팃스+핵심+특징)")
        if "스팃스" not in n and "STIX" not in n.upper():
            problems.append("브랜드명 누락")
            fixes.append("앞에 '스팃스' 추가")
    elif "쿠팡" in platform or "로켓" in platform:
        if len(n) > 50:
            problems.append(f"쿠팡 50자 초과({len(n)}자)")
            fixes.append("핵심키워드+사이즈+용도만 남기기")
        if len(n) < 20:
            problems.append("키워드 부족")
            fixes.append("DIY+사이즈+초보자 추가")
    else:
        if len(n) < 40:
            problems.append("지마켓/11번가 키워드 부족")
            fixes.append("유사키워드·카테고리어 추가")

    if not re.search(r"\d+x\d+", n, re.I) and any(k in n for k in ("보석", "십자", "액자", "캔버스")):
        if "40" not in n and "30" not in n and "50" not in n and "25" not in n:
            problems.append("사이즈 미표기")
            fixes.append("40x50 또는 30x40 표기")

    if "보석" not in n and "십자" not in n and "DIY" not in n and "접착" not in n and "본드" not in n:
        if any(k in n for k in ("도안", "키트", "만들기")):
            problems.append("핵심카테고리키워드 누락")
            fixes.append("보석십자수/DIY 키워드 앞쪽 배치")

    if re.search(r"40x50cm|30x40cm", n, re.I):
        problems.append("사이즈 형식")
        fixes.append("40x50 (cm 제거, 쿠팡용)")

    if not problems:
        return "C", "양호", "-"

    pri = "A" if len(problems) >= 2 or "50자" in str(problems) else "B"
    return pri, " / ".join(problems), " / ".join(fixes)


def load_platform(name: str, path: Path):
    if not path.exists():
        return [], []
    if path.suffix.lower() == ".csv":
        return read_csv(path)
    return read_xlsx(path)


def main():
    summary = []
    all_issues = []

    for platform, path in FILES.items():
        headers, rows = load_platform(platform, path)
        names = []
        for row in rows:
            nm = find_name_col(headers, row)
            if nm:
                names.append(nm)
        summary.append({"platform": platform, "total": len(rows), "with_name": len(names), "file": path.name})

        for nm in names:
            pri, prob, fix = analyze_name(nm, platform)
            if pri != "C":
                all_issues.append({
                    "priority": pri, "platform": platform, "name": nm,
                    "problem": prob, "fix": fix,
                })

    all_issues.sort(key=lambda x: (0 if x["priority"] == "A" else 1, x["platform"], x["name"]))

    wb = openpyxl.Workbook()
    hf = PatternFill("solid", fgColor="7A5C46")
    hfont = Font(bold=True, color="FFFFFF")

    ws = wb.active
    ws.title = "플랫폼_요약"
    ws.append(["플랫폼", "전체건수", "상품명추출", "파일"])
    for c in ws[1]:
        c.font, c.fill = hfont, hf
    for s in summary:
        ws.append([s["platform"], s["total"], s["with_name"], s["file"]])

    ws2 = wb.create_sheet("개선안_전체")
    ws2.append(["우선순위", "플랫폼", "현재상품명", "문제점", "개선안"])
    for c in ws2[1]:
        c.font, c.fill = hfont, hf
    for item in all_issues:
        ws2.append([item["priority"], item["platform"], item["name"], item["problem"], item["fix"]])

    ws3 = wb.create_sheet("우선순위A")
    ws3.append(["플랫폼", "현재상품명", "문제점", "개선안"])
    for c in ws3[1]:
        c.font, c.fill = hfont, hf
    for item in all_issues:
        if item["priority"] == "A":
            ws3.append([item["platform"], item["name"], item["problem"], item["fix"]])

    # 플랫폼별 상품명 샘플
    ws4 = wb.create_sheet("상품명_전체")
    ws4.append(["플랫폼", "상품명"])
    for c in ws4[1]:
        c.font, c.fill = hfont, hf
    for platform, path in FILES.items():
        _, rows = load_platform(platform, path)
        headers, _ = load_platform(platform, path)
        for row in rows:
            nm = find_name_col(headers, row)
            if nm:
                ws4.append([platform, nm])

    for sh in wb.worksheets:
        for col in sh.columns:
            sh.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 2, 80
            )

    out = OUT / f"STIX_MD_전체상품분석_{TODAY}.xlsx"
    wb.save(out)

    # txt report
    lines = [
        f"STIX MD 전체상품 분석 {TODAY}",
        "=" * 50,
        "",
        "[플랫폼별 상품 수]",
    ]
    for s in summary:
        lines.append(f"  {s['platform']}: {s['total']:,}건 ({s['file']})")
    lines += ["", f"[개선 필요] A={sum(1 for x in all_issues if x['priority']=='A')}건, B={sum(1 for x in all_issues if x['priority']=='B')}건", ""]
    lines.append("[우선순위 A TOP 15]")
    for item in [x for x in all_issues if x["priority"] == "A"][:15]:
        lines.append(f"  [{item['platform']}] {item['name'][:60]}")
        lines.append(f"    문제: {item['problem']}")
        lines.append(f"    개선: {item['fix']}")
    (OUT / f"MD_전체상품분석_{TODAY}.txt").write_text("\n".join(lines), encoding="utf-8")
    print(f"saved {out}")
    for s in summary:
        print(f"  {s['platform']}: {s['total']} products, issues A={sum(1 for x in all_issues if x['platform']==s['platform'] and x['priority']=='A')}")


if __name__ == "__main__":
    main()
