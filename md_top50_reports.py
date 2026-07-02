# -*- coding: utf-8 -*-
"""STIX MD - TOP50 키워드 / 상세·썸네일 개선 50"""
import csv
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

SRC = Path(__file__).parent / "쇼핑몰별 전체상품"
OUT = Path(__file__).parent
TODAY = datetime.now().strftime("%Y-%m-%d")

PLATFORMS = {
    "쿠팡": ("coupang", SRC / "쿠팡.xlsx"),
    "로켓그로스": ("rocket", SRC / "로켓배송.로켓그로스.xlsx"),
    "지마켓/옥션": ("gmarket", SRC / "지마켓,옥션.xlsx"),
    "스마트스토어": ("smartstore", SRC / "스마트스토어.csv"),
    "11번가": ("11st", SRC / "11번가.xlsx"),
    "카페24": ("cafe24", SRC / "카페24.자사몰.xlsx"),
}

HF = PatternFill("solid", fgColor="7A5C46")
HFONT = Font(bold=True, color="FFFFFF")


def clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def read_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    hdr_i = 0
    for i, row in enumerate(rows[:10]):
        cells = [clean(c) for c in row if c]
        if any("등록상품명" in c or c == "상품명" for c in cells):
            hdr_i = i
            break
    headers = [clean(h) for h in rows[hdr_i]]
    out = []
    for row in rows[hdr_i + 1 :]:
        if not any(row):
            continue
        d = {headers[j]: row[j] for j in range(min(len(headers), len(row))) if headers[j]}
        if d:
            out.append(d)
    return out


def read_csv(path: Path) -> list[dict]:
    with open(path, encoding="utf-8-sig", newline="") as f:
        return [dict(r) for r in csv.DictReader(f) if any(r.values())]


def get_field(row: dict, *keys) -> str:
    for k in keys:
        for hk, hv in row.items():
            if k in hk and hv not in (None, ""):
                return clean(hv)
    return ""


def get_name(row: dict) -> str:
    return get_field(row, "등록상품명", "상품명", "쿠팡 노출상품명") or ""


def get_id(row: dict, platform: str) -> str:
    return get_field(
        row,
        "등록상품ID", "상품번호", "마스터상품번호", "SKU ID",
        "상품번호(스마트스토어)", "상품코드", "판매자상품코드",
    )


def extract_size(name: str) -> str:
    m = re.search(r"(\d+)\s*[xX×]\s*(\d+)\s*(?:cm)?", name, re.I)
    if m:
        return f"{m.group(1)}x{m.group(2)}"
    return ""


def design_core(name: str) -> str:
    n = name
    for p in (
        r"^\(.*?\)\s*", r"^일본.*?[\s/]", r"^신규 상품 프로모션 대상\s*",
        r"^스팃스\s*", r"^보석십자수\s*", r"^new\s*", r"^\[.*?\]\s*",
    ):
        n = re.sub(p, "", n, flags=re.I)
    n = re.sub(r"\s*(DIY|키트|캔버스|액자형|판매자배송|로켓그로스).*$", "", n, flags=re.I)
    n = re.sub(r"\d+x\d+.*$", "", n, flags=re.I)
    return clean(n)[:25] or "인기도안"


def category_of(name: str) -> str:
    n = name.lower()
    if any(k in n for k in ("b7000", "e6000", "b6000", "접착", "본드", "치약본드")):
        return "접착제"
    if "보석십자수" in n or "비즈십자" in n:
        return "보석십자수"
    if "어린이" in n and ("스티커" in n or "보석" in n):
        return "어린이보석십자수"
    if "십자수" in n or "자수" in n:
        return "십자수"
    if "일본" in n or "서적" in n:
        return "일본서적"
    if "마크라" in n or "부자재" in n:
        return "부자재"
    return "기타"


def sales_score(name: str, row: dict, platform: str) -> float:
    cat = category_of(name)
    score = 0.0
    n = name.lower()
    # 카테고리 가중 (쿠팡 MD 핵심)
    score += {"접착제": 400, "보석십자수": 350, "어린이보석십자수": 320, "십자수": 200, "부자재": 150}.get(cat, 50)
    if "일본" in n and cat == "일본서적":
        score += 80  # 서적도 판매 있음
    if "스팃스" in name:
        score += 80
    if extract_size(name):
        score += 40
    if "40x50" in n or "30x40" in n:
        score += 30
    if "해바라기" in n or "어린이" in n:
        score += 25
    # 가격
    price = 0
    for k in ("판매가", "판매가격", "상품가"):
        v = get_field(row, k).replace(",", "")
        if v.isdigit():
            price = int(v)
            break
    if 12000 <= price <= 45000:
        score += 40
    elif 8000 <= price <= 60000:
        score += 20
    # 재고 (낮을수록 회전 추정 - 약한 신호)
    stock = get_field(row, "재고수량", "재고")
    if stock.isdigit():
        s = int(stock)
        if s < 500:
            score += 30
        elif s < 5000:
            score += 15
    # 판매중
    st = get_field(row, "판매상태", "판매상태")
    if "판매" in st and "중지" not in st:
        score += 20
    # 로켓/쿠팡 핵심 SKU
    if any(k in n for k in ("b7000", "e6000")):
        score += 200
    if platform == "로켓그로스" and cat in ("접착제", "보석십자수"):
        score += 100
    return score


def seo_name(name: str, platform: str) -> tuple[str, str]:
    """returns (optimized_name, recommended_tags)"""
    cat = category_of(name)
    size = extract_size(name) or ("30x40" if "30" in name else "40x50")
    core = design_core(name)
    tags = []

    if cat == "접착제":
        prod = "E6000" if "e6000" in name.lower() else "B7000" if "b7000" in name.lower() else "B6000" if "b6000" in name.lower() else "접착제"
        ml = re.search(r"(\d+)\s*ml", name, re.I)
        vol = f" {ml.group(1)}ml" if ml else ""
        if platform == "쿠팡":
            opt = f"{prod} 투명 접착제{vol} 치약본드 공예 DIY".strip()[:50]
        elif platform in ("지마켓/옥션",):
            opt = f"{prod} 접착제{vol} 치약본드 공예용 다용도 본드 DIY 스팃스 만들기 키트".strip()[:80]
        elif platform == "스마트스토어":
            opt = f"스팃스 {prod} 투명 접착제{vol} 공예 DIY 치약본드".strip()
        elif platform == "11번가":
            opt = f"{prod} 접착제{vol} DIY 공예 추천 치약본드 초보자".strip()[:60]
        else:
            opt = f"스팃스 {prod} 접착제{vol} 공예 DIY 치약본드 초보자용".strip()
        tags = [prod, "접착제", "치약본드", "공예용품", "DIY", "스팃스"]
    elif cat == "보석십자수":
        if platform == "쿠팡":
            opt = f"보석십자수 캔버스 {size} {core} DIY 키트 초보자"[:50]
        elif platform == "지마켓/옥션":
            opt = f"보석십자수 DIY 키트 캔버스형 {size} {core} 액자형 취미 만들기 집콕 비즈십자수"[:80]
        elif platform == "스마트스토어":
            opt = f"스팃스 보석십자수 DIY 키트 액자형 {size} {core} 취미 만들기"
        elif platform == "11번가":
            opt = f"보석십자수 DIY 키트 {size} {core} 액자형 추천 초보자"[:60]
        else:
            opt = f"보석십자수 {size} DIY 키트 {core} 초보자도 쉽게 완성하는 취미세트"
        tags = ["보석십자수", "DIY키트", f"보석십자수{size}", "캔버스보석십자수", "취미키트", "집콕취미"]
    elif cat == "어린이보석십자수":
        if platform == "쿠팡":
            opt = "어린이 보석십자수 스티커 DIY 12종 만들기 키트"[:50]
        elif platform == "지마켓/옥션":
            opt = "어린이 보석십자수 스티커 DIY 만들기 키트 12종 놀이세트 캐릭터"[:80]
        elif platform == "스마트스토어":
            opt = "스팃스 어린이 보석십자수 스티커 만들기 DIY 12종 세트"
        else:
            opt = f"어린이 보석십자수 스티커 DIY 키트 {core} 만들기 추천"
        tags = ["어린이보석십자수", "스티커DIY", "키즈취미", "만들기"]
    elif cat == "십자수":
        if platform == "쿠팡":
            opt = f"십자수 패키지 {core} DIY 자수세트"[:50]
        elif platform == "지마켓/옥션":
            opt = f"십자수 패키지 {core} DIY 프랑스자수 취미 자수세트 만들기"[:80]
        elif platform == "스마트스토어":
            opt = f"스팃스 십자수 패키지 {core} DIY 자수 키트"
        else:
            opt = f"십자수 패키지 {core} DIY 자수 키트 취미"
        tags = ["십자수패키지", "DIY자수", "프랑스자수", "자수세트"]
    else:
        opt = name[:50] if platform == "쿠팡" else name[:80]
        tags = ["DIY", "수예", "스팃스"]

    return clean(opt), ",".join(tags)


def detail_score(item: dict) -> float:
    name = item["name"]
    score = 0.0
    issues = []
    cat = category_of(name)
    if cat == "일본서적":
        score += 40
        issues.append("일본서적-한글구성/FAQ 필요")
    if not extract_size(name) and cat in ("보석십자수", "십자수"):
        score += 35
        issues.append("사이즈표/FAQ 없음(상품명에 사이즈 없음)")
    if cat == "보석십자수":
        score += 30
        issues.append("4분할 HTML 미적용 추정")
    if cat == "접착제":
        score += 25
        issues.append("안전정보/용량/사용법 섹션 필요")
    if "스팃스" not in name and item["platform"] == "스마트스토어":
        score += 20
        issues.append("브랜드 스토리/신뢰 섹션 부족")
    reg = item.get("reg_date", "")
    if reg and reg[:4] <= "2024":
        score += 15
        issues.append("구형 상세(2024 이전 등록)")
    if len(name) < 25:
        score += 10
        issues.append("상품명 짧음-상세 보완 필요")
    item["detail_issues"] = issues
    item["detail_fix"] = (
        "1)히어로:완성작+한줄카피 → 2)구성품4컷 → 3)작업과정 → 4)사이즈표+FAQ"
        if cat in ("보석십자수", "어린이보석십자수")
        else "1)핵심특징 → 2)구성/스펙 → 3)사용법 → 4)FAQ+배송"
    )
    return score


def thumb_score(item: dict) -> float:
    name = item["name"]
    score = 0.0
    issues = []
    if not extract_size(name) and category_of(name) == "보석십자수":
        score += 40
        issues.append("사이즈 배지 없음")
    img = item.get("image", "")
    if img.lower().endswith(".png"):
        score += 25
        issues.append("PNG 썸네일(밝은 JPG 권장)")
    if category_of(name) == "보석십자수" and not any(k in name for k in ("액자", "캔버스", "완성")):
        score += 30
        issues.append("완성작+구성품 4컷 미구분")
    if "스팃스" not in name and item["platform"] in ("스마트스토어", "11번가"):
        score += 15
        issues.append("브랜드/클릭 유도 약함")
    reg = item.get("reg_date", "")
    if reg and reg[:4] <= "2024":
        score += 20
        issues.append("구형 썸네일 추정")
    if category_of(name) == "접착제" and "ml" not in name.lower():
        score += 15
        issues.append("용량/제품 클로즈업 없음")
    item["thumb_issues"] = issues
    item["thumb_fix"] = (
        "001_thumb:완성작+사이즈배지 / 002_main:밝은배경 / 003_detail:구성품 / 004_finish:인테리어"
        if category_of(name) in ("보석십자수", "어린이보석십자수")
        else "001:제품전체 / 002:클로즈업 / 003:구성/스펙 / 004:사용장면"
    )
    return score


def load_all() -> dict[str, list[dict]]:
    data = {}
    for plat, (_, path) in PLATFORMS.items():
        if not path.exists():
            continue
        rows = read_csv(path) if path.suffix.lower() == ".csv" else read_xlsx(path)
        items = []
        for row in rows:
            name = get_name(row)
            if not name or name.startswith("Catalog"):
                continue
            items.append({
                "platform": plat,
                "name": name,
                "id": get_id(row, plat),
                "price": get_field(row, "판매가", "판매가격", "상품가"),
                "search_kw": get_field(row, "검색어", "검색어설정"),
                "image": get_field(row, "대표이미지 URL", "이미지등록(상세)"),
                "reg_date": get_field(row, "상품등록일", "최초등록일", "최종수정일", "최종변경일")[:10],
                "row": row,
            })
        data[plat] = items
    return data


def style_header(ws):
    for c in ws[1]:
        c.font, c.fill = HFONT, HF


def auto_width(ws, mx=70):
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 2, mx
        )


def main():
    all_data = load_all()
    all_items = [it for items in all_data.values() for it in items]

    wb = openpyxl.Workbook()

    # ── 1. 플랫폼별 TOP50 키워드 ──
    ws1 = wb.active
    ws1.title = "1_TOP50_키워드재조합"
    ws1.append([
        "플랫폼", "순위", "상품ID", "현재상품명", "개선상품명", "추천태그/검색어",
        "선정점수", "카테고리", "승인",
    ])
    style_header(ws1)
    for plat, items in all_data.items():
        ranked = sorted(items, key=lambda x: sales_score(x["name"], x["row"], plat), reverse=True)[:50]
        for i, it in enumerate(ranked, 1):
            opt, tags = seo_name(it["name"], plat)
            sc = sales_score(it["name"], it["row"], plat)
            ws1.append([
                plat, i, it["id"], it["name"], opt, tags,
                round(sc, 1), category_of(it["name"]), "",
            ])

    # ── 2. 상세페이지 개선 50 ──
    ws2 = wb.create_sheet("2_상세페이지_개선50")
    ws2.append(["순위", "플랫폼", "상품ID", "상품명", "문제점", "개선 HTML 구조", "승인"])
    style_header(ws2)
    for it in all_items:
        detail_score(it)
    top_detail = sorted(all_items, key=detail_score, reverse=True)[:50]
    for i, it in enumerate(top_detail, 1):
        ws2.append([
            i, it["platform"], it["id"], it["name"],
            " / ".join(it.get("detail_issues", [])),
            it.get("detail_fix", ""), "",
        ])

    # ── 3. 썸네일 개선 50 ──
    ws3 = wb.create_sheet("3_썸네일_개선50")
    ws3.append(["순위", "플랫폼", "상품ID", "상품명", "문제점", "4컷 구조 제안", "승인"])
    style_header(ws3)
    for it in all_items:
        thumb_score(it)
    top_thumb = sorted(all_items, key=thumb_score, reverse=True)[:50]
    for i, it in enumerate(top_thumb, 1):
        ws3.append([
            i, it["platform"], it["id"], it["name"],
            " / ".join(it.get("thumb_issues", [])),
            it.get("thumb_fix", ""), "",
        ])

    for sh in wb.worksheets:
        auto_width(sh)

    out = OUT / f"STIX_MD_TOP50_작업목록_{TODAY}.xlsx"
    wb.save(out)

    # txt 요약
    lines = [
        f"STIX MD TOP50 작업목록 {TODAY}",
        f"총 상품 DB: {len(all_items):,}건",
        "",
        "[1] 플랫폼별 TOP50 키워드 재조합 → 시트1 (플랫폼당 50건)",
        "[2] 상세페이지 개선 50건 → 시트2",
        "[3] 썸네일 개선 50건 → 시트3",
        "",
        "※ 판매량 컬럼 없음 → 카테고리·가격·재고·핵심SKU 가중치로 상위 추정",
        "",
        "=== 플랫폼별 TOP5 키워드 샘플 ===",
    ]
    for plat, items in all_data.items():
        ranked = sorted(items, key=lambda x: sales_score(x["name"], x["row"], plat), reverse=True)[:5]
        lines.append(f"\n[{plat}]")
        for it in ranked:
            opt, _ = seo_name(it["name"], plat)
            lines.append(f"  · {it['name'][:45]}")
            lines.append(f"    → {opt}")

    (OUT / f"MD_TOP50_작업목록_{TODAY}.txt").write_text("\n".join(lines), encoding="utf-8")
    print(f"saved {out}")
    print(f"products {len(all_items)}, sheets: top50 x {len(all_data)}, detail50, thumb50")


if __name__ == "__main__":
    main()
