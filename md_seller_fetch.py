# -*- coding: utf-8 -*-
"""STIX MD - 쇼핑몰 MD Chrome(9233) CDP 연결 수집"""
import json
import re
import shutil
import sys
import time
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from playwright.sync_api import sync_playwright

OUT_DIR = Path(__file__).parent
TODAY = datetime.now().strftime("%Y-%m-%d")
SHOT_DIR = OUT_DIR / "screenshots"
SHOT_DIR.mkdir(exist_ok=True)
CDP_PORTS = [9233, 9222]
DEFAULT_UD = Path.home() / "AppData/Local/Google/Chrome/User Data"
MD_PROFILE = OUT_DIR / "chrome_md_profile"

TARGETS = [
    ("coupang", "쿠팡 Wing", [
        "https://wing.coupang.com/tenants/seller-web/vendor-inventory/list",
        "https://wing.coupang.com/vendor-inventory/list",
    ]),
    ("esm", "지마켓/옥션", [
        "https://www.esmplus.com/Home/v2/goods-manage",
    ]),
    ("smartstore", "스마트스토어", [
        "https://sell.smartstore.naver.com/#/products/origin-list",
    ]),
    ("11st", "11번가", [
        "https://soffice.11st.co.kr/view/main",
    ]),
    ("cafe24", "카페24 스팃스", [
        "https://escall.cafe24.com/disp/admin/shop1/main/dashboard",
    ]),
]

SEARCH_KWS = ["보석십자수", "스팃스 보석십자수", "십자수 패키지"]

PLATFORM_DOMAIN = {
    "coupang": "wing.coupang",
    "esm": "esmplus.com",
    "smartstore": "smartstore.naver.com",
    "11st": "soffice.11st",
    "cafe24": "cafe24.com",
}


def clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def cdp_alive(port: int) -> bool:
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:{port}/json/version", timeout=2) as r:
            return bool(r.read())
    except Exception:
        return False


def find_cdp_port() -> int | None:
    for p in CDP_PORTS:
        if cdp_alive(p):
            return p
    return None


def sync_profile_copy():
    """Default 프로필에서 로그인 세션 파일 복사 (Chrome 종료 후)"""
    src = DEFAULT_UD / "Default"
    dst = MD_PROFILE / "Default"
    dst.mkdir(parents=True, exist_ok=True)
    # User Data 루트 파일
    for fn in ["Local State"]:
        s, d = DEFAULT_UD / fn, MD_PROFILE / fn
        if s.exists():
            try:
                shutil.copy2(s, d)
            except Exception:
                pass
    files = [
        "Cookies", "Login Data", "Preferences", "Web Data",
        "Network Persistent State", "Secure Preferences",
    ]
    for fn in files:
        s, d = src / fn, dst / fn
        if s.is_dir():
            try:
                if d.exists():
                    shutil.rmtree(d)
                shutil.copytree(s, d, dirs_exist_ok=True)
            except Exception:
                pass
        elif s.exists():
            try:
                shutil.copy2(s, d)
            except Exception:
                pass
    for sub in ["Local Storage", "IndexedDB", "Sessions"]:
        ls_src, ls_dst = src / sub, dst / sub
        if ls_src.exists():
            try:
                if ls_dst.exists():
                    shutil.rmtree(ls_dst)
                shutil.copytree(ls_src, ls_dst, dirs_exist_ok=True)
            except Exception:
                pass


def launch_md_chrome(port: int = 9233):
    import subprocess
    sync_profile_copy()
    chrome = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    user_data = str(MD_PROFILE)
    args = [
        chrome,
        f"--remote-debugging-port={port}",
        f"--user-data-dir={user_data}",
        "--profile-directory=Default",
        "--no-first-run",
        "https://wing.coupang.com/",
    ]
    subprocess.Popen(args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, close_fds=True)
    for _ in range(40):
        time.sleep(1)
        if cdp_alive(port):
            return port
    return None


def is_logged_in(page) -> bool:
    u = page.url.lower()
    if any(x in u for x in ("login", "signin", "xauth", "auth/realms", "nidlogin")):
        return False
    try:
        body = page.locator("body").inner_text(timeout=3000)[:800]
        if any(x in body for x in (
            "다시 한번 확인해주세요", "로그인 하기", "판매자 로그인",
            "자동 로그아웃", "로그인이 필요",
        )):
            return False
        if "404" in body and "찾을 수 없" in body:
            return False
    except Exception:
        pass
    return True


def dismiss_popups(page):
    for sel in (
        'button:has-text("확인")', 'button:has-text("닫기")',
        'button:has-text("close")', '[aria-label="Close"]', '.btn_close',
    ):
        try:
            loc = page.locator(sel)
            for i in range(min(loc.count(), 3)):
                btn = loc.nth(i)
                if btn.is_visible():
                    btn.click(timeout=2000)
                    page.wait_for_timeout(400)
        except Exception:
            pass
    try:
        cb = page.locator('label:has-text("오늘")').first
        if cb.count() and cb.is_visible():
            cb.click(timeout=2000)
    except Exception:
        pass


def on_platform(page, key: str) -> bool:
    domain = PLATFORM_DOMAIN.get(key, "")
    return bool(domain and domain in page.url.lower())


def navigate_product_list(page, key: str) -> bool:
    """로그인된 탭에서 상품목록 화면으로 이동. False = 세션 만료 등."""
    if not on_platform(page, key):
        return False
    dismiss_popups(page)
    try:
        if key == "smartstore":
            if "#/products/origin-list" not in page.url:
                page.evaluate("() => { window.location.hash = '#/products/origin-list'; }")
                page.wait_for_timeout(3000)
            if "#/products/origin-list" not in page.url:
                page.goto(
                    "https://sell.smartstore.naver.com/#/products/origin-list",
                    timeout=90000, wait_until="domcontentloaded",
                )
            page.wait_for_timeout(10000)
        elif key == "esm":
            if "goods-manage" not in page.url:
                page.goto(
                    "https://www.esmplus.com/Home/v2/goods-manage",
                    timeout=90000, wait_until="domcontentloaded",
                )
            page.wait_for_timeout(8000)
            dismiss_popups(page)
            body = page.locator("body").inner_text(timeout=3000)[:600]
            if "자동 로그아웃" in body:
                return False
            # 상품목록은 item.esmplus.com/goods/list iframe 안에 있음
            target = page
            for fr in page.frames:
                if "item.esmplus.com/goods/list" in (fr.url or ""):
                    target = fr
                    break
            clicked = False
            loc = target.locator('button:has-text("검색")')
            for i in range(loc.count()):
                try:
                    el = loc.nth(i)
                    if el.is_visible() and clean(el.inner_text(timeout=800)) == "검색":
                        el.click(timeout=8000)
                        clicked = True
                        break
                except Exception:
                    pass
            if not clicked:
                try:
                    target.get_by_role("button", name="검색", exact=True).click(timeout=8000)
                except Exception:
                    pass
            for _ in range(30):
                try:
                    fbody = target.locator("body").inner_text(timeout=3000)
                    if re.search(r"상품목록\s*[1-9]", fbody):
                        break
                except Exception:
                    pass
                page.wait_for_timeout(1000)
        elif key == "11st":
            on_list = bool(re.search(r"view/\d+|prd|product|SellProduct", page.url, re.I))
            if not on_list or "찾을 수 없" in page.title():
                page.goto(
                    "https://soffice.11st.co.kr/view/main",
                    timeout=90000, wait_until="domcontentloaded",
                )
                page.wait_for_timeout(5000)
                for sel in ('text=상품관리', 'a:has-text("상품관리")', 'span:has-text("상품관리")'):
                    try:
                        menu = page.locator(sel).first
                        if menu.count() and menu.is_visible():
                            menu.click(timeout=5000)
                            page.wait_for_timeout(2000)
                            break
                    except Exception:
                        pass
                for sel in (
                    'text=상품조회/수정', 'a:has-text("상품조회")',
                    'span:has-text("상품조회")', 'text=상품 조회/수정',
                ):
                    try:
                        link = page.locator(sel).first
                        if link.count() and link.is_visible():
                            link.click(timeout=5000)
                            page.wait_for_timeout(10000)
                            break
                    except Exception:
                        pass
            page.wait_for_timeout(5000)
            for sel in ('button:has-text("검색")', 'button:has-text("조회")'):
                try:
                    btn = page.locator(sel).first
                    if btn.count() and btn.is_visible():
                        btn.click(timeout=3000)
                        page.wait_for_timeout(12000)
                        break
                except Exception:
                    pass
            page.evaluate("window.scrollTo(0, 800)")
            page.wait_for_timeout(2000)
        elif key == "cafe24":
            dismiss_popups(page)
            if "product" not in page.url.lower() or "다시 한번" in page.locator("body").inner_text(timeout=2000)[:200]:
                page.goto(
                    "https://escall.cafe24.com/disp/admin/shop1/main/dashboard",
                    timeout=90000, wait_until="domcontentloaded",
                )
                page.wait_for_timeout(5000)
                dismiss_popups(page)
                for sel in ('text=상품목록 전체보기', 'a:has-text("상품목록")', 'text=상품 목록'):
                    try:
                        link = page.locator(sel).first
                        if link.count() and link.is_visible():
                            link.click(timeout=5000)
                            page.wait_for_timeout(10000)
                            break
                    except Exception:
                        pass
            page.wait_for_timeout(5000)
        page.evaluate("window.scrollTo(0, document.body.scrollHeight / 3)")
        page.wait_for_timeout(1500)
    except Exception:
        pass
    return is_logged_in(page)


def extract_products(page, platform_key: str = "") -> list[str]:
    found = []
    # SPA 로딩 대기
    page.wait_for_timeout(3000)

    if platform_key == "smartstore":
        for sel in (
            "table tbody tr td a",
            "[class*='product'] a",
            "table tbody tr td:nth-child(8)",
            "table tbody tr td:nth-child(4)",
        ):
            loc = page.locator(sel)
            for i in range(min(loc.count(), 80)):
                try:
                    el = loc.nth(i)
                    t = clean(el.inner_text(timeout=1500))
                    if 8 < len(t) < 200 and t not in found and not t.isdigit():
                        if re.match(r"^\d{4}\.\d{2}\.\d{2}", t):
                            continue
                        if any(x in t for x in ("수정", "복사", "전시", "판매중", "상세설명")):
                            continue
                        found.append(t)
                except Exception:
                    pass
            if len(found) >= 10:
                break

    elif platform_key == "11st":
        target = page
        for fr in page.frames:
            if "SellProductAction" in (fr.url or ""):
                target = fr
                break
        for sel in (
            'input[type="button"][value="검색"]',
            'input[type="submit"][value="검색"]',
            'button:has-text("검색")',
        ):
            try:
                loc = target.locator(sel)
                for i in range(loc.count() - 1, -1, -1):
                    btn = loc.nth(i)
                    if btn.is_visible():
                        btn.click(timeout=5000)
                        target.wait_for_timeout(12000)
                        break
            except Exception:
                pass
        try:
            body = target.locator("body").inner_text(timeout=8000)
            for line in body.split("\n"):
                t = clean(line)
                if 12 < len(t) < 200 and t not in found:
                    if any(k in t for k in ("스팃", "보석", "십자", "DIY", "접착", "E6000", "B6000", "공예", "키트")):
                        if not any(x in t for x in (">", "홈/취미", "검색", "전체", "엑셀", "카테고리")):
                            found.append(t)
        except Exception:
            pass

    elif platform_key == "esm":
        target = page
        for fr in page.frames:
            if "item.esmplus.com/goods/list" in (fr.url or ""):
                target = fr
                break
        try:
            body = target.locator("body").inner_text(timeout=8000)
            for line in body.split("\n"):
                t = clean(line)
                if 12 < len(t) < 200 and t not in found:
                    if any(k in t for k in ("스팃", "보석", "십자", "DIY", "접착", "E6000", "B6000", "공예", "키트", "자수", "원단", "크로바", "와펜")):
                        # 카테고리 경로(>) / 안내문 제외
                        if ">" in t or any(x in t for x in ("검색", "전체", "엑셀", "카테고리", "하위 분류", "더보기")):
                            continue
                        found.append(t)
        except Exception:
            pass

    elif platform_key == "cafe24":
        for sel in ["a[href*='Product']", "td.product", "table tbody tr td", ".product_name"]:
            loc = page.locator(sel)
            for i in range(min(loc.count(), 60)):
                try:
                    t = clean(loc.nth(i).inner_text(timeout=1500))
                    if 8 < len(t) < 200 and t not in found and not t.startswith("0 원"):
                        if any(k in t for k in ("보석", "십자", "스팃", "DIY", "접착")):
                            found.append(t)
                except Exception:
                    pass
            if len(found) >= 5:
                break

    if not found:
        js = """() => {
            const out = [];
            document.querySelectorAll('td, a, span, div, input, p, li').forEach(el => {
                let t = (el.value || el.innerText || el.title || '').trim().replace(/\\s+/g, ' ');
                if (t.length > 12 && t.length < 180) out.push(t);
            });
            return [...new Set(out)];
        }"""
        for fr in page.frames:
            try:
                for t in fr.evaluate(js):
                    t = clean(t)
                    if t not in found and any(k in t for k in ("보석", "십자", "스팃", "DIY", "접착", "본드", "패키지", "E6000", "B6000", "공예", "키트", "액자", "자수", "마크라")):
                        if not any(x in t for x in ("로그인", "로그아웃", "전체", "선택삭제", "매뉴얼", "FAQ", "Open API", "광고안내", "셀러존", "카카오톡", "예약 관리", "여행", "숙박")):
                            found.append(t)
            except Exception:
                pass
        if not found:
            try:
                for t in page.evaluate(js):
                    t = clean(t)
                    if t not in found and any(k in t for k in ("보석", "십자", "스팃", "DIY", "접착", "본드", "패키지", "E6000", "B6000", "공예", "키트", "액자", "자수", "마크라")):
                        if not any(x in t for x in ("로그인", "로그아웃", "전체", "선택삭제", "매뉴얼", "FAQ", "Open API", "광고안내", "셀러존", "카카오톡")):
                            found.append(t)
            except Exception:
                pass
        if not found:
            try:
                body = page.locator("body").inner_text(timeout=5000)
                for line in body.split("\n"):
                    t = clean(line)
                    if 10 < len(t) < 180 and t not in found:
                        if any(k in t for k in ("보석", "십자", "스팃", "DIY", "접착", "본드", "패키지", "E6000", "B6000", "공예", "키트")):
                            if not any(x in t for x in ("로그인", "로그아웃", "저장", "선택", "전체", "건", "매뉴얼", "FAQ")):
                                found.append(t)
            except Exception:
                pass
    return found[:50]


def fetch_target(page, key, name, urls) -> dict:
    rec = {"name": name, "logged_in": False, "url": "", "products": [], "rows": []}
    safe_name = name.replace("/", "_")
    for url in urls:
        try:
            if not on_platform(page, key):
                page.goto(url, timeout=90000, wait_until="domcontentloaded")
                page.wait_for_timeout(7000 if key != "coupang" else 5000)
            elif key == "coupang":
                page.goto(url, timeout=90000, wait_until="domcontentloaded")
                page.wait_for_timeout(7000)
            if not on_platform(page, key):
                rec["url"] = page.url
                continue
            if key != "coupang":
                ok = navigate_product_list(page, key)
                rec["logged_in"] = ok
            else:
                page.wait_for_timeout(3000)
                rec["logged_in"] = is_logged_in(page)
            rec["url"] = page.url
            if rec["logged_in"]:
                rec["products"] = extract_products(page, key)
                if rec["products"] or key == "coupang":
                    break
        except Exception as e:
            rec["error"] = str(e)
    try:
        rec["screenshot"] = str(SHOT_DIR / f"{safe_name}.png")
        page.screenshot(path=rec["screenshot"])
    except Exception:
        pass
    return rec


def coupang_search(page) -> list[dict]:
    out = []
    for kw in SEARCH_KWS:
        try:
            page.goto(f"https://www.coupang.com/np/search?q={kw}", timeout=60000)
            page.wait_for_timeout(4000)
            if "denied" in page.title().lower():
                continue
            cards = page.locator("li.search-product")
            for i in range(min(cards.count(), 20)):
                c = cards.nth(i)
                try:
                    title = clean(c.locator(".name").first.inner_text(timeout=2000))
                    price = ""
                    if c.locator(".price-value").count():
                        price = clean(c.locator(".price-value").first.inner_text(timeout=1500))
                    ours = any(b in title for b in ("스팃스", "STIX", "스팟스"))
                    out.append({"kw": kw, "rank": i + 1, "title": title, "price": price, "ours": ours})
                except Exception:
                    pass
        except Exception:
            pass
    return out


def write_report(platforms, search):
    wb = Workbook()
    hf, hfont = PatternFill("solid", fgColor="7A5C46"), Font(bold=True, color="FFFFFF")

    ws = wb.active
    ws.title = "접속결과"
    ws.append(["플랫폼", "로그인", "URL", "추출수"])
    for c in ws[1]:
        c.font, c.fill = hfont, hf
    for p in platforms:
        ws.append([p["name"], "Y" if p["logged_in"] else "N", p.get("url", ""), len(p["products"])])

    ws2 = wb.create_sheet("상품_데이터")
    ws2.append(["플랫폼", "내용"])
    for c in ws2[1]:
        c.font, c.fill = hfont, hf
    for p in platforms:
        for x in p["products"]:
            ws2.append([p["name"], x])

    ws3 = wb.create_sheet("쿠팡검색")
    ws3.append(["키워드", "순위", "상품명", "가격", "우리"])
    for c in ws3[1]:
        c.font, c.fill = hfont, hf
    for s in search:
        ws3.append([s["kw"], s["rank"], s["title"], s["price"], "Y" if s["ours"] else ""])

    for sh in wb.worksheets:
        for col in sh.columns:
            sh.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 2, 90)

    xlsx = OUT_DIR / f"STIX_MD_실데이터_{TODAY}.xlsx"
    wb.save(xlsx)
    (OUT_DIR / f"STIX_MD_실데이터_{TODAY}.json").write_text(
        json.dumps({"platforms": platforms, "search": search}, ensure_ascii=False, indent=2), encoding="utf-8")
    return xlsx


def main():
    port = find_cdp_port()
    if port == 9222:
        print("port 9222 = blog profile, launching MD chrome on 9233...")
        import subprocess
        subprocess.run(["taskkill", "/F", "/IM", "chrome.exe"], capture_output=True)
        time.sleep(3)
        port = launch_md_chrome(9233)
    elif not port:
        print("no cdp, sync profile + launch MD chrome on 9233...")
        import subprocess
        subprocess.run(["taskkill", "/F", "/IM", "chrome.exe"], capture_output=True)
        time.sleep(3)
        port = launch_md_chrome(9233)
    if not port:
        print("FAIL: run start_chrome_for_md.bat then retry")
        return 1

    print(f"CDP port {port}")
    platforms = []
    search = []
    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(f"http://127.0.0.1:{port}")
        ctx = browser.contexts[0]

        def get_page_for(key: str):
            domain = PLATFORM_DOMAIN.get(key, "")
            for pg in ctx.pages:
                if domain and domain in pg.url.lower():
                    return pg
            return None

        for key, name, urls in TARGETS:
            print(f"  {name}...")
            existing = get_page_for(key)
            if existing:
                target_page = existing
            else:
                target_page = ctx.new_page()
            platforms.append(fetch_target(target_page, key, name, urls))
            time.sleep(1)
        search = []
        try:
            wing = get_page_for("coupang")
            if wing:
                search = coupang_search(wing)
        except Exception:
            pass

    path = write_report(platforms, search)
    print(f"saved {path}")
    for pl in platforms:
        print(f"  {pl['name']}: login={pl['logged_in']} n={len(pl['products'])}")
    if not any(p["logged_in"] and p["products"] for p in platforms):
        print("fallback: md_history_fetch...")
        import subprocess
        subprocess.run([sys.executable, str(OUT_DIR / "md_history_fetch.py")], check=False)
    return 0


if __name__ == "__main__":
    sys.exit(main())
