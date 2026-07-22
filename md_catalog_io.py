# -*- coding: utf-8 -*-
"""STIX MD - 쇼핑몰 export 공통 로더 (상품명·가격·재고·판매량)"""
from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent
SRC = BASE / "쇼핑몰별 전체상품"
SALES_SRC = BASE / "쇼핑몰별 매출주문"


def _find_catalog(*patterns: str) -> Path:
    """로컬 export 파일명 변형(공백·날짜 접미) 자동 매칭."""
    for pat in patterns:
        hits = sorted(SRC.glob(pat))
        if hits:
            return hits[0]
    return SRC / patterns[-1]


PLATFORM_FILES: dict[str, Path] = {
    "쿠팡": _find_catalog("쿠팡*전체상품*.xlsx", "쿠팡.xlsx"),
    "로켓그로스": _find_catalog("로켓배송*로켓그로스*.xlsx", "로켓배송.로켓그로스.xlsx"),
    "지마켓/옥션": _find_catalog("지마켓,옥션.xlsx"),
    "스마트스토어": _find_catalog("스마트스토어*전체상품*.csv", "스마트스토어.csv"),
    "11번가": _find_catalog("11번가*전체상품*.xlsx", "11번가.xlsx"),
    "카페24": _find_catalog("카페24*자사몰*.xlsx", "카페24.자사몰.xlsx"),
}

NAME_KEYS = (
    "등록상품명", "쿠팡 노출상품명", "상품명", "노출상품명",
    "판매자상품명", "스마트스토어전용 상품명", "상품 이름", "ProductName",
)
PRICE_KEYS = (
    "판매가", "판매가격", "상품가", "할인가", "노출가격", "판매가(원)",
    "판매자할인가", "최종판매가", "가격",
)
STOCK_KEYS = ("재고수량", "재고", "재고량", "옵션재고", "수량")
STATUS_KEYS = ("판매상태", "판매상태명", "상태", "전시상태", "노출상태")
SALES_KEYS = (
    "판매량", "7일판매량", "30일판매량", "누적판매량", "주문수", "결제건수",
    "판매수량", "최근7일판매", "최근30일판매", "아이템위너 판매량",
)
ID_KEYS = (
    "등록상품ID", "상품번호", "마스터상품번호", "SKU ID", "옵션 ID",
    "상품번호(스마트스토어)", "상품코드", "판매자상품코드", "vendorItemId",
)


def clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def parse_int(val) -> int | None:
    if val is None or val == "":
        return None
    s = clean(val).replace(",", "").replace("원", "")
    if s.lstrip("-").isdigit():
        return int(s)
    m = re.search(r"-?\d+", s)
    return int(m.group()) if m else None


def parse_price(val) -> int | None:
    return parse_int(val)


def get_field(row: dict, *keys: str) -> str:
    for k in keys:
        for hk, hv in row.items():
            if hk and k in hk and hv not in (None, ""):
                return clean(hv)
    return ""


def _pick_xlsx_sheet(wb):
    """Prefer data sheets (e.g. Coupang Template) over Help/안내."""
    preferred = ("Template", "상품", "전체상품", "Sheet1", "시트1")
    for name in preferred:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        low = name.lower()
        if low in {"help", "hidden", "안내"}:
            continue
        return wb[name]
    return wb.active


def read_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = _pick_xlsx_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    hdr_i = 0
    best_score = -1
    for i, row in enumerate(rows[:40]):
        cells = [clean(c) for c in row if c]
        score = sum(
            1
            for c in cells
            for kw in (
                "등록상품명",
                "쿠팡 노출상품명",
                "상품명",
                "ProductName",
                "판매가",
                "재고",
                "등록상품ID",
            )
            if kw in c
        )
        # skip long help banners
        if any(len(c) > 80 for c in cells[:2]):
            continue
        if score > best_score:
            best_score = score
            hdr_i = i
    headers = [clean(h) for h in rows[hdr_i]]
    out = []
    for row in rows[hdr_i + 1 :]:
        if not any(row):
            continue
        d = {headers[j]: row[j] for j in range(min(len(headers), len(row))) if headers[j]}
        if d:
            out.append(d)
    return out


def read_csv(path: Path) -> list[dict]:
    with open(path, encoding="utf-8-sig", newline="") as f:
        return [dict(r) for r in csv.DictReader(f) if any(r.values())]


def load_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    if path.suffix.lower() == ".csv":
        return read_csv(path)
    return read_xlsx(path)


def normalize_row(row: dict, platform: str) -> dict:
    name = get_field(row, *NAME_KEYS)
    price = parse_price(get_field(row, *PRICE_KEYS))
    if price is None:
        for k in PRICE_KEYS:
            price = parse_price(row.get(k))
            if price is not None:
                break
    stock = parse_int(get_field(row, *STOCK_KEYS))
    if stock is None:
        for k in STOCK_KEYS:
            stock = parse_int(row.get(k))
            if stock is not None:
                break
    sales = parse_int(get_field(row, *SALES_KEYS))
    if sales is None:
        for k in SALES_KEYS:
            sales = parse_int(row.get(k))
            if sales is not None:
                break
    status = get_field(row, *STATUS_KEYS)
    pid = get_field(row, *ID_KEYS)
    return {
        "platform": platform,
        "id": pid,
        "name": name,
        "price": price,
        "stock": stock,
        "sales": sales,
        "status": status,
        "raw": row,
    }


def load_all_catalogs() -> list[dict]:
    items = []
    for platform, path in PLATFORM_FILES.items():
        for row in load_rows(path):
            rec = normalize_row(row, platform)
            if rec["name"]:
                items.append(rec)
    return items


def load_sales_exports() -> dict[str, list[dict]]:
    """쇼핑몰별 매출주문/*.xlsx|csv — 파일명에 기간 포함 권장 (예: 쿠팡_2026-06-01_2026-06-30.xlsx)"""
    out: dict[str, list[dict]] = {}
    if not SALES_SRC.exists():
        return out
    for path in sorted(SALES_SRC.glob("*")):
        if path.suffix.lower() not in (".xlsx", ".csv", ".xls"):
            continue
        plat = path.stem.split("_")[0]
        rows = load_rows(path)
        out.setdefault(plat, []).extend(
            normalize_row(r, plat) for r in rows if get_field(r, *NAME_KEYS) or get_field(r, *ID_KEYS)
        )
    return out
