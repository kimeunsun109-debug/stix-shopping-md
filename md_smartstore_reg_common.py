# -*- coding: utf-8 -*-
"""스마트스토어 등록정보 보완 — 공통 유틸."""
from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

BASE = Path(__file__).parent
DATA_DIR = BASE / "쇼핑몰별 전체상품"
CSV_PATH = DATA_DIR / "스마트스토어  전체상품목록.csv"

CDP = "http://127.0.0.1:9233"

# 30건 테스트 배치 — 이전 세션 태그 반영 성공 ID
TAG_BATCH_SUCCESS_IDS = {
    "13623403164", "13622502531", "13599815751", "13599623268", "13599552309",
    "13365720299", "13021569547", "12668524257", "12668517912", "12668512904",
    "12650594339", "12601107115", "12596537354",
}


def tag_batch_failed_ids(limit: int = 30) -> list[str]:
    rows = load_selling_rows(find_excel()) if find_excel().exists() else []
    if not rows:
        rows = [r for r in load_csv_rows() if r["status"] == "판매중"]
    batch = [r["product_id"] for r in rows[:limit]]
    return [pid for pid in batch if pid not in TAG_BATCH_SUCCESS_IDS]

REAL_BRANDS = {"크로바", "DMC", "튤립"}
STORE_REJECT_TAGS = {"스팃스", "STIX", "stix", "스마트스토어", "크로바", "DMC", "튤립"}

CATEGORY_TAG_POOL: dict[str, list[str]] = {
    "보석십자수": [
        "십자수", "DIY", "취미", "집콕취미", "비즈아트", "인테리어소품", "취미생활",
        "선물추천", "DIY키트", "액자만들기", "보석십자수", "만들기", "홈데코",
    ],
    "자수": [
        "자수", "십자수", "자수키트", "프랑스자수", "수예", "DIY", "취미",
        "자수실", "자수바늘", "자수공예", "자수도안", "자수서적", "도안책",
    ],
    "뜨개": [
        "뜨개질", "코바늘", "대바늘", "모사용", "뜨개바늘", "뜨개실", "수예",
        "DIY", "취미", "뜨개질바늘", "뜨개가방", "뜨개목도리", "일본뜨개서적",
    ],
    "부자재": [
        "수예용품", "부자재", "공예용품", "DIY", "취미", "만들기", "수예",
        "공예재료", "수예도구", "핸드폰고리", "키링", "연결고리", "소품diy",
    ],
    "접착": ["접착제", "본드", "E6000", "B7000", "공예용접착", "DIY", "수예", "만들기"],
    "default": ["DIY", "취미", "수예", "만들기", "취미생활", "선물추천", "홈데코", "인테리어", "집콕취미"],
}


def find_excel() -> Path:
    candidates = sorted(BASE.glob("STIX_스마트스토어_등록정보보완_*.xlsx"), reverse=True)
    if candidates:
        return candidates[0]
    out = BASE / f"STIX_스마트스토어_등록정보보완_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return out


def resolve_brand_mfr(name: str, csv_brand: str = "", csv_mfr: str = "") -> tuple[str, str]:
    text = f"{name} {csv_brand} {csv_mfr}"
    for b in REAL_BRANDS:
        if b.lower() in text.lower() or b in text:
            return b, b
    return "스팃스", "스팃스"


def detect_pool(name: str, category: str = "") -> list[str]:
    text = f"{name} {category}"
    if re.search(r"보석십자수|비즈", text):
        return CATEGORY_TAG_POOL["보석십자수"]
    if re.search(r"자수서적|도안집|자수책|자수도안|프랑스자수|십자수편", text):
        return CATEGORY_TAG_POOL["자수"]
    if re.search(r"뜨개|코바늘|대바늘|단수", text):
        return CATEGORY_TAG_POOL["뜨개"]
    if re.search(r"핸드폰|키링|스트랩|부자재|연결|고리", text):
        return CATEGORY_TAG_POOL["부자재"]
    if re.search(r"E6000|B7000|접착", text):
        return CATEGORY_TAG_POOL["접착"]
    if re.search(r"자수|십자수", text):
        return CATEGORY_TAG_POOL["자수"]
    return CATEGORY_TAG_POOL["default"]


def extract_name_tags(name: str, limit: int = 12) -> list[str]:
    """상품명에서 상품특화 태그 후보 추출."""
    tags: list[str] = []
    patterns = [
        (r"(\d+x\d+)", lambda m: m.group(1).replace(" ", "")),
        (r"(보석십자수)", lambda m: m.group(1)),
        (r"(프랑스자수|십자수|자수실|자수도안|자수서적)", lambda m: m.group(1)),
        (r"(뜨개가방|뜨개질|대바늘|코바늘)", lambda m: m.group(1)),
        (r"(핸드폰|키링|스트랩|오거나이저|단수계)", lambda m: m.group(1)),
        (r"(DIY|초보자|액자형|판넬)", lambda m: m.group(1)),
    ]
    for pat, fn in patterns:
        m = re.search(pat, name, re.I)
        if m:
            t = fn(m)
            if t and t not in tags:
                tags.append(t)
    # 도안명 일부
    for chunk in re.split(r"[\s,/]+", name):
        chunk = re.sub(r"[^\w가-힣]", "", chunk)
        if len(chunk) >= 3 and chunk not in {"스팃스", "일본서적", "NEW", "New"}:
            compound = f"{chunk}보석십자수" if "보석" not in chunk and "십자" in name else chunk
            if 3 <= len(compound) <= 20 and compound not in tags:
                tags.append(compound)
    return tags[:limit]


def build_tags(name: str, category: str = "", brand: str = "") -> str:
    pool = detect_pool(name, category)
    specific = extract_name_tags(name)
    merged: list[str] = []
    for t in specific + pool:
        t = t.strip()
        if not t or t in STORE_REJECT_TAGS:
            continue
        if brand and t == brand:
            continue
        if t not in merged:
            merged.append(t)
        if len(merged) >= 18:
            break
    return ", ".join(merged[:18])


def filter_tags_for_store(tags: list[str], brand: str = "") -> list[str]:
    out: list[str] = []
    skip = STORE_REJECT_TAGS | {brand} if brand else STORE_REJECT_TAGS
    for t in tags:
        t = re.sub(r"\s*\(\d+\)\s*$", "", t.strip())
        if not t or t in skip:
            continue
        if t not in out:
            out.append(t)
        if len(out) >= 10:
            break
    return out


def load_csv_rows() -> list[dict]:
    rows: list[dict] = []
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            pid = (r.get("상품번호(스마트스토어)") or "").strip()
            if not pid:
                continue
            name = (r.get("상품명") or "").strip()
            brand, mfr = resolve_brand_mfr(name, r.get("브랜드명", ""), r.get("제조사명", ""))
            cat = " ".join(
                x for x in [r.get("대분류"), r.get("중분류"), r.get("소분류"), r.get("세분류")] if x
            )
            rows.append(
                {
                    "product_id": pid,
                    "name": name,
                    "status": (r.get("판매상태") or "").strip(),
                    "display": (r.get("전시상태") or "").strip(),
                    "brand": brand,
                    "manufacturer": mfr,
                    "category": cat,
                    "tags": build_tags(name, cat, brand),
                    "seller_code": (r.get("판매자상품코드") or "").strip(),
                    "modified": (r.get("최종수정일") or "").strip(),
                }
            )
    return rows


def load_selling_rows(excel_path: Path | None = None) -> list[dict]:
    path = excel_path or find_excel()
    if path.exists():
        wb = load_workbook(path, read_only=True, data_only=True)
        if "판매중" in wb.sheetnames:
            ws = wb["판매중"]
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {h: i for i, h in enumerate(headers)}
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                pid = str(row[idx.get("상품번호", 0)] or "").strip()
                if not pid:
                    continue
                rows.append(
                    {
                        "product_id": pid,
                        "name": str(row[idx.get("상품명", 1)] or ""),
                        "brand": str(row[idx.get("브랜드", 2)] or ""),
                        "manufacturer": str(row[idx.get("제조사", 3)] or ""),
                        "tags": str(row[idx.get("SEO태그", 4)] or ""),
                        "category": str(row[idx.get("카테고리", 5)] or ""),
                    }
                )
            wb.close()
            return rows
        wb.close()
    return [r for r in load_csv_rows() if r["status"] == "판매중"]


def write_review_excel(path: Path | None = None) -> Path:
    out = path or find_excel()
    selling = [r for r in load_csv_rows() if r["status"] == "판매중"]
    wb = Workbook()
    ws = wb.active
    ws.title = "판매중"
    ws.append(["상품번호", "상품명", "브랜드", "제조사", "SEO태그", "카테고리", "판매상태", "전시상태"])
    for r in selling:
        ws.append([
            r["product_id"], r["name"], r["brand"], r["manufacturer"],
            r["tags"], r["category"], r["status"], r["display"],
        ])
    wb.save(out)
    return out
