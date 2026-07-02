# -*- coding: utf-8 -*-
"""STIX MD - 전체상품 가격·재고·매출기회/주의 분석 + (선택) 기간별 매출 비교"""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from md_catalog_io import (
    PLATFORM_FILES,
    SALES_SRC,
    load_all_catalogs,
    load_rows,
    normalize_row,
    parse_int,
)

OUT = Path(__file__).parent
TODAY = datetime.now().strftime("%Y-%m-%d")
HF = PatternFill("solid", fgColor="7A5C46")
HFONT = Font(bold=True, color="FFFFFF")


def category_hint(name: str) -> str:
    n = name.lower()
    if any(k in n for k in ("b7000", "e6000", "b6000", "접착", "본드")):
        return "접착제"
    if "보석" in n or "드릴" in n:
        return "보석십자수"
    if "십자" in n:
        return "십자수"
    if "어린이" in n or "스티커" in n:
        return "어린이"
    if any(k in n for k in ("dmc", "크로바", "바늘", "원단", "부자재")):
        return "부자재"
    return "기타"


def is_selling(status: str) -> bool:
    s = status or ""
    return any(x in s for x in ("판매중", "전시중", "노출", "Selling")) and "중지" not in s and "금지" not in s


def is_stopped(status: str) -> bool:
    s = status or ""
    return any(x in s for x in ("판매중지", "판매금지", "품절", "판매종료", "중지"))


def opportunity_score(rec: dict) -> float:
    """매출 올릴 수 있는 상품 추정 점수"""
    score = 0.0
    name = rec["name"]
    cat = category_hint(name)
    score += {"접착제": 400, "보석십자수": 350, "어린이": 280, "십자수": 200, "부자재": 120}.get(cat, 50)
    if "스팃스" in name or "STIX" in name.upper():
        score += 60
    p = rec.get("price")
    if p and 8000 <= p <= 45000:
        score += 40
    st = rec.get("stock")
    if st is not None:
        if 1 <= st <= 500:
            score += 35  # 회전 중 추정
        elif st >= 50000:
            score -= 20  # 과다재고
    if is_selling(rec.get("status", "")):
        score += 30
    else:
        score -= 50
    if rec.get("sales"):
        score += min(rec["sales"] * 2, 200)
    if not re.search(r"\d+x\d+", name, re.I) and cat in ("보석십자수", "십자수"):
        score += 25  # SEO 개선 여지
    return score


def risk_score(rec: dict) -> float:
    """매출 떨어질/손실 위험 점수"""
    score = 0.0
    if is_stopped(rec.get("status", "")):
        score += 80
    if rec.get("price") == 0:
        score += 70
    st = rec.get("stock")
    if st == 0:
        score += 60
    elif st and st >= 99999:
        score += 25
    if rec.get("sales") == 0:
        score += 30
    cat = category_hint(rec["name"])
    if cat in ("접착제", "보석십자수") and is_stopped(rec.get("status", "")):
        score += 40
    return score


def price_issues(rec: dict) -> list[str]:
    issues = []
    p = rec.get("price")
    if p is None:
        issues.append("가격 미표기")
    elif p <= 0:
        issues.append("가격 0원")
    elif p < 1000:
        issues.append("비정상 저가")
    elif p > 200000:
        issues.append("고가(전환율 점검)")
    name = rec["name"].lower()
    if "보석" in name or "십자" in name:
        if p and p < 5000:
            issues.append("카테고리 대비 저가")
        if p and p > 80000:
            issues.append("카테고리 대비 고가")
    return issues


def compare_sales_periods() -> list[dict]:
    """쇼핑몰별 매출주문/ 폴더에 기간별 export 2개 이상이면 SKU별 판매량 변화"""
    if not SALES_SRC.exists():
        return []
    files = sorted(SALES_SRC.glob("*"))
    if len(files) < 2:
        return []

    by_file: list[tuple[str, dict[str, int]]] = []
    for path in files[-2:]:
        rows = load_rows(path)
        agg: dict[str, int] = defaultdict(int)
        for row in rows:
            rec = normalize_row(row, path.stem.split("_")[0])
            key = rec["id"] or rec["name"]
            if not key:
                continue
            qty = rec.get("sales") or parse_int(
                next((row[k] for k in row if "수량" in k or "건수" in k or "판매" in k), None)
            )
            if qty:
                agg[key] += qty
        by_file.append((path.name, dict(agg)))

    if len(by_file) < 2:
        return []
    prev_name, prev = by_file[0]
    curr_name, curr = by_file[1]
    changes = []
    all_keys = set(prev) | set(curr)
    for key in all_keys:
        p, c = prev.get(key, 0), curr.get(key, 0)
        if p == 0 and c == 0:
            continue
        delta = c - p
        pct = (delta / p * 100) if p else (100 if c else 0)
        changes.append({
            "key": key,
            "prev": p,
            "curr": c,
            "delta": delta,
            "pct": round(pct, 1),
            "prev_file": prev_name,
            "curr_file": curr_name,
            "trend": "하락" if delta < 0 else ("상승" if delta > 0 else "유지"),
        })
    changes.sort(key=lambda x: x["delta"])
    return changes


def style_header(ws):
    for c in ws[1]:
        c.font, c.fill = HFONT, HF


def main():
    items = load_all_catalogs()
    sales_changes = compare_sales_periods()

    wb = openpyxl.Workbook()

    # 요약
    ws = wb.active
    ws.title = "플랫폼_요약"
    ws.append(["플랫폼", "파일", "건수", "판매중", "품절/중지", "가격있음", "판매량컬럼"])
    style_header(ws)
    plat_stats = defaultdict(lambda: {"total": 0, "selling": 0, "stopped": 0, "priced": 0, "with_sales": 0})
    for it in items:
        p = it["platform"]
        plat_stats[p]["total"] += 1
        if is_selling(it.get("status", "")):
            plat_stats[p]["selling"] += 1
        if is_stopped(it.get("status", "")):
            plat_stats[p]["stopped"] += 1
        if it.get("price"):
            plat_stats[p]["priced"] += 1
        if it.get("sales") is not None:
            plat_stats[p]["with_sales"] += 1
    for platform, path in PLATFORM_FILES.items():
        st = plat_stats[platform]
        exists = "Y" if path.exists() else "N(없음)"
        ws.append([
            platform, path.name if path.exists() else exists,
            st["total"], st["selling"], st["stopped"], st["priced"], st["with_sales"],
        ])

    # 전체 카탈로그
    ws2 = wb.create_sheet("전체_카탈로그")
    ws2.append(["플랫폼", "상품ID", "상품명", "판매가", "재고", "판매량", "판매상태", "카테고리"])
    style_header(ws2)
    for it in sorted(items, key=lambda x: (x["platform"], x["name"])):
        ws2.append([
            it["platform"], it["id"], it["name"], it.get("price"), it.get("stock"),
            it.get("sales"), it.get("status"), category_hint(it["name"]),
        ])

    # 가격 이상
    ws3 = wb.create_sheet("가격_이상")
    ws3.append(["플랫폼", "상품명", "판매가", "문제", "권장"])
    style_header(ws3)
    for it in items:
        issues = price_issues(it)
        if issues:
            fix = "가격 재확인·경쟁가 조사" if "가격" in issues[0] else "정상가 설정"
            ws3.append([it["platform"], it["name"], it.get("price"), " / ".join(issues), fix])

    # 재고 위험
    ws4 = wb.create_sheet("재고_위험")
    ws4.append(["플랫폼", "상품명", "재고", "판매가", "유형", "권장"])
    style_header(ws4)
    for it in items:
        st = it.get("stock")
        if st is None:
            continue
        note, action = "", ""
        if st == 0 and is_selling(it.get("status", "")):
            note, action = "품절위험(판매중)", "재고 보충 또는 판매중지"
        elif 1 <= st <= 3:
            note, action = "재고 임박", "긴급 발주"
        elif st >= 50000:
            note, action = "과다재고", "할인·광고·로켓회수 검토"
        if note:
            ws4.append([it["platform"], it["name"], st, it.get("price"), note, action])

    # 매출 올릴 SKU TOP50
    ws5 = wb.create_sheet("매출_기회_TOP50")
    ws5.append(["순위", "플랫폼", "상품명", "판매가", "재고", "판매량", "점수", "액션"])
    style_header(ws5)
    ranked_up = sorted(items, key=opportunity_score, reverse=True)[:50]
    for i, it in enumerate(ranked_up, 1):
        actions = []
        if not re.search(r"\d+x\d+", it["name"], re.I) and category_hint(it["name"]) in ("보석십자수", "십자수"):
            actions.append("사이즈·SEO")
        if it.get("stock") and it["stock"] >= 50000:
            actions.append("할인테스트")
        if category_hint(it["name"]) == "접착제":
            actions.append("가격·아이템위너")
        if not actions:
            actions.append("썸네일·상세")
        ws5.append([
            i, it["platform"], it["name"], it.get("price"), it.get("stock"),
            it.get("sales"), round(opportunity_score(it), 1), " / ".join(actions),
        ])

    # 매출 주의 TOP50
    ws6 = wb.create_sheet("매출_주의_TOP50")
    ws6.append(["순위", "플랫폼", "상품명", "판매가", "재고", "상태", "점수", "액션"])
    style_header(ws6)
    ranked_down = sorted(items, key=risk_score, reverse=True)[:50]
    for i, it in enumerate(ranked_down, 1):
        act = []
        if is_stopped(it.get("status", "")):
            act.append("재활성 또는 정리")
        if it.get("stock") == 0:
            act.append("품절 해소")
        if it.get("price") == 0:
            act.append("가격 입력")
        if not act:
            act.append("모니터링")
        ws6.append([
            i, it["platform"], it["name"], it.get("price"), it.get("stock"),
            it.get("status"), round(risk_score(it), 1), " / ".join(act),
        ])

    # 기간별 매출 변화 (파일 2개 이상일 때)
    ws7 = wb.create_sheet("매출_기간비교")
    ws7.append(["SKU/상품", "이전", "현재", "증감", "증감%", "추세", "이전파일", "현재파일"])
    style_header(ws7)
    if sales_changes:
        for ch in sales_changes[:200]:
            ws7.append([
                ch["key"], ch["prev"], ch["curr"], ch["delta"], ch["pct"],
                ch["trend"], ch["prev_file"], ch["curr_file"],
            ])
    else:
        ws7.append([
            "(데이터 없음)", "", "", "", "",
            f"'{SALES_SRC.name}' 폴더에 기간별 주문/매출 export 2개 이상 필요",
            "", "",
        ])

    for sh in wb.worksheets:
        for col in sh.columns:
            sh.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 2, 85
            )

    xlsx = OUT / f"STIX_MD_가격매출분석_{TODAY}.xlsx"
    wb.save(xlsx)

    # TXT 요약
    lines = [
        f"STIX MD 가격·매출 분석 {TODAY}",
        "=" * 55,
        "",
        "[플랫폼별]",
    ]
    for platform, path in PLATFORM_FILES.items():
        st = plat_stats[platform]
        flag = "✓" if path.exists() else "✗"
        lines.append(
            f"  {flag} {platform}: {st['total']:,}건 | 판매중 {st['selling']:,} | "
            f"중지/품절 {st['stopped']:,} | export={path.name if path.exists() else '없음'}"
        )
    lines += [
        "",
        "[매출 기회 TOP10]",
    ]
    for i, it in enumerate(ranked_up[:10], 1):
        lines.append(f"  {i}. [{it['platform']}] {it['name'][:55]} | {it.get('price')}원")
    lines += ["", "[매출 주의 TOP10]"]
    for i, it in enumerate(ranked_down[:10], 1):
        lines.append(f"  {i}. [{it['platform']}] {it['name'][:55]} | {it.get('status')}")
    if sales_changes:
        lines += ["", "[매출 하락 TOP5 (기간비교)]"]
        for ch in sales_changes[:5]:
            lines.append(f"  {ch['key'][:40]} | {ch['prev']}→{ch['curr']} ({ch['delta']:+d})")
    else:
        lines += [
            "",
            f"[매출 추세] '{SALES_SRC.name}' 폴더에 주문/매출 export 2개 넣으면 하락/상승 SKU 자동 비교",
        ]
    txt = OUT / f"MD_가격매출분석_{TODAY}.txt"
    txt.write_text("\n".join(lines), encoding="utf-8")

    print(f"saved {xlsx}")
    print(f"saved {txt}")
    print(f"  total SKUs: {len(items):,}")
    print(f"  platforms with data: {sum(1 for p in PLATFORM_FILES.values() if p.exists())}/{len(PLATFORM_FILES)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
