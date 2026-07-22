# -*- coding: utf-8 -*-
"""Mode B — manual collector (xlsx/csv/txt/json/url paste)."""
from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path

from seo_engine.collectors.base import BaseCollector
from seo_engine.models import CollectionBundle, ProductSnapshot


def _parse_price(v) -> int | None:
    if v is None:
        return None
    s = re.sub(r"[^\d]", "", str(v))
    return int(s) if s else None


def _row_to_product(row: dict, rank: int | None = None) -> ProductSnapshot:
    def g(*keys: str) -> str:
        for k in keys:
            for rk, rv in row.items():
                if str(rk).strip().lower() == k.lower() and rv not in (None, ""):
                    return str(rv).strip()
        return ""

    reviews_raw = g("reviews", "리뷰", "review_text", "후기")
    reviews = [x.strip() for x in re.split(r"[\n|;]", reviews_raw) if x.strip()]
    detail = g("detail", "상세", "detail_text", "상세페이지")
    images = g("image", "images", "image_urls", "대표이미지")
    image_urls = [x.strip() for x in re.split(r"[,|\n]", images) if x.strip()]
    rating_s = g("rating", "평점")
    try:
        rating = float(rating_s) if rating_s else None
    except ValueError:
        rating = None
    rc = g("review_count", "리뷰수", "reviews_count")
    try:
        review_count = int(re.sub(r"[^\d]", "", rc)) if rc else None
    except ValueError:
        review_count = None

    return ProductSnapshot(
        title=g("title", "상품명", "name", "노출상품명"),
        brand=g("brand", "브랜드"),
        price=_parse_price(g("price", "가격", "판매가")),
        review_count=review_count,
        rating=rating,
        image_urls=image_urls,
        detail_text=detail,
        detail_bullets=[x.strip() for x in detail.split("\n") if x.strip()][:30],
        reviews=reviews,
        url=g("url", "상품URL", "link"),
        product_id=g("product_id", "상품번호", "id"),
        rank=rank,
        category=g("category", "카테고리"),
        option_names=[
            x.strip()
            for x in re.split(r"[,|;]", g("options", "옵션", "option_names"))
            if x.strip()
        ],
        image_notes=g("image_notes", "이미지메모", "image_note"),
        raw={str(k): v for k, v in row.items()},
    )


class ManualCollector(BaseCollector):
    """
    Inputs (any combination):
      - keyword
      - mine: ProductSnapshot fields or dict
      - competitors: list[dict]
      - path: .xlsx / .csv / .json / .txt
        sheets/rows: role=mine|competitor, title, brand, price, ...
    """

    def __init__(
        self,
        *,
        keyword: str = "",
        mine: ProductSnapshot | dict | None = None,
        competitors: list[ProductSnapshot | dict] | None = None,
        path: str | Path | None = None,
    ) -> None:
        self.keyword = keyword
        self.mine = mine
        self.competitors = competitors or []
        self.path = Path(path) if path else None

    def collect(self) -> CollectionBundle:
        mine: ProductSnapshot | None = None
        comps: list[ProductSnapshot] = []
        keyword = self.keyword

        if self.path and self.path.exists():
            loaded_kw, loaded_mine, loaded_comps = self._load_file(self.path)
            keyword = keyword or loaded_kw
            mine = loaded_mine
            comps = loaded_comps

        if isinstance(self.mine, ProductSnapshot):
            mine = self.mine
        elif isinstance(self.mine, dict):
            mine = _row_to_product(self.mine)

        for c in self.competitors:
            if isinstance(c, ProductSnapshot):
                comps.append(c)
            elif isinstance(c, dict):
                comps.append(_row_to_product(c, rank=c.get("rank")))

        if mine is None:
            raise ValueError("Mode B: 내 상품(mine) 정보가 필요합니다.")
        if not mine.title:
            raise ValueError("Mode B: 내 상품명이 비어 있습니다.")

        return CollectionBundle(
            keyword=keyword or self._infer_keyword(mine.title),
            mine=mine,
            competitors=comps[:10],
            source="manual",
            marketplace="coupang",
            collected_at=datetime.now().isoformat(timespec="seconds"),
        )

    def _infer_keyword(self, title: str) -> str:
        for k in ("보석십자수", "E6000", "B7000", "십자수", "DIY"):
            if k.lower() in title.lower() or k in title:
                return k
        return title.split()[0] if title.split() else "검색어"

    def _load_file(self, path: Path) -> tuple[str, ProductSnapshot | None, list[ProductSnapshot]]:
        suf = path.suffix.lower()
        if suf == ".json":
            data = json.loads(path.read_text(encoding="utf-8"))
            return self._from_json(data)
        if suf == ".csv":
            with path.open(encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))
            return self._from_rows(rows)
        if suf in {".xlsx", ".xlsm"}:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                if any(v not in (None, "") for v in d.values()):
                    rows.append(d)
            wb.close()
            return self._from_rows(rows)
        if suf == ".txt":
            return self._from_txt(path.read_text(encoding="utf-8"))
        raise ValueError(f"지원하지 않는 파일 형식: {suf}")

    def _from_json(self, data: dict) -> tuple[str, ProductSnapshot | None, list[ProductSnapshot]]:
        kw = str(data.get("keyword") or data.get("검색어") or "")
        mine = data.get("mine") or data.get("내상품")
        comps = data.get("competitors") or data.get("경쟁") or []
        mine_p = _row_to_product(mine) if isinstance(mine, dict) else None
        if isinstance(mine, str):
            mine_p = ProductSnapshot(title=mine)
        comp_ps = []
        for i, c in enumerate(comps, 1):
            if isinstance(c, str):
                comp_ps.append(ProductSnapshot(title=c, rank=i))
            elif isinstance(c, dict):
                comp_ps.append(_row_to_product(c, rank=c.get("rank") or i))
        return kw, mine_p, comp_ps

    def _from_rows(self, rows: list[dict]) -> tuple[str, ProductSnapshot | None, list[ProductSnapshot]]:
        kw = ""
        mine = None
        comps: list[ProductSnapshot] = []
        for i, row in enumerate(rows, 1):
            role = str(row.get("role") or row.get("구분") or row.get("type") or "").lower()
            if not kw:
                kw = str(row.get("keyword") or row.get("검색어") or "")
            p = _row_to_product(row, rank=i if "comp" in role or role in {"경쟁", "competitor"} else None)
            if role in {"mine", "내상품", "me", "자사"}:
                mine = p
            elif role in {"competitor", "경쟁", "comp", "상위"}:
                comps.append(p)
            elif mine is None:
                mine = p
            else:
                comps.append(p)
        return kw, mine, comps

    def _from_txt(self, text: str) -> tuple[str, ProductSnapshot | None, list[ProductSnapshot]]:
        """
        간단한 텍스트 포맷:
          검색어: 보석십자수
          내상품: ...
          경쟁1: ...
          리뷰: ...
        """
        kw = ""
        mine_title = ""
        comps: list[str] = []
        reviews: list[str] = []
        detail = ""
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            if ":" in line:
                k, v = line.split(":", 1)
                k, v = k.strip(), v.strip()
                if k in {"검색어", "keyword"}:
                    kw = v
                elif k in {"내상품", "mine", "상품명"}:
                    mine_title = v
                elif k.startswith("경쟁") or k.lower().startswith("comp"):
                    comps.append(v)
                elif k in {"리뷰", "review"}:
                    reviews.append(v)
                elif k in {"상세", "detail"}:
                    detail = v
            elif not mine_title:
                mine_title = line
            else:
                comps.append(line)
        mine = ProductSnapshot(title=mine_title, reviews=reviews, detail_text=detail) if mine_title else None
        return kw, mine, [ProductSnapshot(title=t, rank=i) for i, t in enumerate(comps, 1)]
