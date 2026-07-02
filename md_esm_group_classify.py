# -*- coding: utf-8 -*-
"""지마켓/옥션 그룹관리 업로드용 상품 분류"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent
SRC = BASE / "쇼핑몰별 전체상품" / "지마켓,옥션.xlsx"
TODAY = datetime.now().strftime("%Y-%m-%d")
OUT_XLSX = BASE / f"STIX_지마켓_그룹관리_업로드_{TODAY}.xlsx"
OUT_SUMMARY = BASE / f"MD_지마켓_그룹관리_분류요약_{TODAY}.txt"

TEMPLATE_COLS = [
    "마스터상품번호",
    "상품번호",
    "상품명",
    "판매가",
    "판매상태",
    "그룹명칭",
    "카테고리",
]


def clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def parse_price(val) -> float:
    if val is None or val == "":
        return 0.0
    s = clean(val).replace(",", "").replace("원", "")
    try:
        return float(s)
    except ValueError:
        m = re.search(r"[\d.]+", s)
        return float(m.group()) if m else 0.0


def has_any(text: str, *words: str) -> bool:
    t = text.lower()
    return any(w.lower() in t for w in words)


def theme_subgroup(text: str, *, kids_ok: bool = True) -> str:
    if kids_ok and has_any(
        text,
        "어린이",
        "키즈",
        "유아",
        "초등",
        "스티커",
        "12종",
        "24종",
        "32종",
        "집콕",
        "캐릭터 diy",
        "점프보이",
        "공주",
        "요정",
        "영웅",
        "자동차",
    ):
        return "키즈,초급"
    if has_any(
        text,
        "종교",
        "성모",
        "마리아",
        "예수",
        "만찬",
        "부처",
        "반야",
        "기독",
        "천주",
        "불교",
        "명화",
        "모나리자",
        "별이 빛나는",
    ):
        return "종교.명화"
    if has_any(
        text,
        "3개세트",
        "3종세트",
        "2+1",
        "3종",
        "세트 3",
        "3세트",
    ):
        return "3개세트상품"
    if has_any(
        text,
        "꽃",
        "장미",
        "튤립",
        "해바라기",
        "화병",
        "정원",
        "식물",
        "나무",
        "벚꽃",
        "모란",
        "연꽃",
        "잉어",
        "과일",
        "야채",
        "농장",
        "화원",
        "튤립",
        "핑크",
        "장미꽃",
    ):
        return "꽃,식물"
    if has_any(
        text,
        "동물",
        "고양이",
        "강아지",
        "토끼",
        "부엉이",
        "새",
        "나비",
        "물고기",
        "캐릭터",
        "유니콘",
        "곰",
        "펭귄",
        "사슴",
        "호랑이",
        "백호",
        "부엉",
        "강아",
        "고양",
        "토끼",
        "동물",
        "앵무",
        "물개",
    ):
        return "동물,캐릭터"
    if has_any(
        text,
        "여인",
        "소녀",
        "인물",
        "레이디",
        "신부",
        "발레",
        "여자",
        "커플",
        "신랑",
        "웨딩",
        "가족",
    ):
        return "인물감성"
    if has_any(
        text,
        "풍경",
        "바다",
        "해변",
        "항해",
        "저택",
        "마을",
        "파리",
        "골목",
        "산",
        "호수",
        "유럽",
        "성",
        "다리",
        "등대",
        "요트",
        "카페",
        "커피",
        "수중",
        "해변가",
        "풍차",
        "노을",
        "일출",
        "sea view",
        "landscape",
        "뉴욕",
        "런던",
        "시골",
        "가을",
        "겨울",
        "눈",
    ):
        return "풍경,건물"
    return "풍경,건물"


def clover_sub(text: str) -> str:
    if has_any(text, "코바늘", "태팅", "아뮤레", "레이스바늘", "바늘세트", "코바늘"):
        return "코바늘,태팅"
    if has_any(text, "가위", "리퍼", "핑킹", "재단", "트와"):
        return "가위,리퍼"
    if has_any(text, "바늘", "실꿰", "골무", "니팅", "뜨개", "수세미"):
        return "바늘,실꿰기,골무"
    if has_any(text, "요요", "폼폼", "직조", "weaving", "loom"):
        return "요요,폼폼,직조기"
    if has_any(text, "펜", "전사", "먹지", "프린트", "transfer"):
        return "펜,전사지,먹지"
    return "챠코페이퍼,접착제,기타제품"


def basic_sub(text: str) -> str:
    if has_any(text, "코바늘", "태팅"):
        return "코바늘,태팅"
    if has_any(text, "가위", "리퍼", "핑킹"):
        return "가위,리퍼"
    if has_any(text, "바늘", "실꿰", "골무", "니팅"):
        return "바늘,실꿰기,골무"
    if has_any(text, "요요", "폼폼", "실감개", "실 끊", "실끊"):
        return "요요,폼폼,실감개"
    if has_any(text, "펜", "전사", "먹지"):
        return "펜,전사지,먹지"
    return "챠코페이퍼,접착제,기타제품"


def thread_sub(text: str) -> str:
    if has_any(text, "dmc") and has_any(
        text, "25번", "19번", "면사", "자수실", "에코비타", "울사", "태팅", "레이스"
    ):
        return "DMC면사,특수사"
    if has_any(text, "앵커", "anchor"):
        return "앵커,기타자수실"
    if has_any(text, "세트", "50색", "100색", "447색", "505색", "풀세트", "60색", "80색"):
        return "자수실세트"
    return "면사자수실"


def hoop_sub(text: str) -> str:
    if has_any(text, "스탠드", "탁상", "받침대", "거치", "스탠드형"):
        return "스탠드형수틀"
    if has_any(text, "타원", "타원형", "oval"):
        return "타원형수틀"
    if has_any(text, "원형", "원목", "우드", "hoop", "수틀") and not has_any(
        text, "타원"
    ):
        if has_any(text, "clover", "크로바", "57-", "너즈", "turkey", "터키", "prym"):
            return "브랜드수틀"
        return "원형수틀"
    if has_any(text, "clover", "크로바", "57-", "너즈", "prym", "프림"):
        return "브랜드수틀"
    return "원형수틀"


def classify(name: str, opt: str = "", category: str = "") -> tuple[str, str]:
    text = clean(f"{name} {opt} {category}")

    if has_any(text, "b6000", "b7000"):
        if "b7000" in text.lower():
            return "본드", "B7000"
        return "본드", "B6000"

    if has_any(
        text,
        "보석십자수",
        "캔버스",
        "큐빅",
        "비즈",
        "드릴비즈",
        "액자형",
        "캔버스형",
        "캔버스판",
    ):
        if has_any(
            text,
            "코팅",
            "마감제",
            "광택",
            "보빈",
            "비즈함",
            "28칸",
            "케이스",
            "정리함",
            "드릴",
            "펜셋",
            "왁스",
            "부자재",
        ) and not has_any(text, "diy", "키트", "만들기", "40x50", "30x40", "30cm"):
            return "보석십자수 부자재", ""
        sub = theme_subgroup(text)
        if sub == "키즈,초급" or has_any(text, "어린이", "스티커", "12종", "32종"):
            sub = "키즈,초급"
        return "보석십자수 도안형", sub

    if has_any(text, "프린트") and has_any(text, "십자수", "패키지"):
        return "프린트십자수패키지", theme_subgroup(text, kids_ok=False)

    if has_any(text, "패키지") and has_any(text, "십자수") and not has_any(
        text, "프린트", "보석", "캔버스", "큐빅"
    ):
        return "십자수패키지", theme_subgroup(text, kids_ok=False)

    if has_any(text, "홈데코") and has_any(text, "십자수") and has_any(
        text, "도안", "원단", "실"
    ):
        if has_any(text, "프린트"):
            return "프린트십자수패키지", theme_subgroup(text, kids_ok=False)
        return "십자수패키지", theme_subgroup(text, kids_ok=False)

    if has_any(text, "무지원단") and has_any(text, "십자수", "패키지"):
        return "프린트십자수패키지", theme_subgroup(text, kids_ok=False)

    if has_any(
        text,
        "서적",
        "도안집",
        "스티치",
        "stitch",
        "일본서적",
        "일본자수",
        "책",
        "도안서",
        "자수도안",
    ) and has_any(text, "십자수", "자수", "도안"):
        return "도안,서적", ""

    if has_any(text, "튤립", "etimo", "에띠모"):
        return "튤립코바늘", ""

    if has_any(text, "dmc") and not has_any(
        text, "25번", "19번", "면사", "자수실", "에코비타", "울사", "태팅", "레이스", "8m"
    ):
        return "DMC 부자재", ""

    if has_any(text, "clover", "크로바", "57-", "57 ") or re.search(
        r"\b57-\d{3}", text, re.I
    ):
        return "크로바 부자재", clover_sub(text)

    if has_any(text, "양모펠트", "펠트니들", "felting", "펠트 니들"):
        if has_any(text, "바늘", "니들"):
            return "양모펠트", "양모펠트 바늘"
        return "양모펠트", "양모펠트 부자재"

    if has_any(
        text,
        "자수실",
        "면사",
        "프랑스자수",
        "십자수실",
        "dmc",
        "8m",
        "8M",
        "앵커",
        "anchor",
        "태팅",
        "레이스 실",
    ):
        return "자수실", thread_sub(text)

    if has_any(text, "수틀", "자수틀", "hoop", "후프"):
        return "수틀", hoop_sub(text)

    if has_any(text, "ct", "아이다", "aida") or (
        has_any(text, "원단", "천", "fabric") and has_any(text, "십자수", "14ct", "11ct")
    ):
        if has_any(text, "린넨", "면", "20수", "40수", "패브릭", "캔버스", "무지"):
            return "원단", "면,린넨원단"
        return "원단", "십자수 원단"

    if has_any(
        text,
        "정리함",
        "케이스",
        "보관함",
        "보빈",
        "다용도",
        "수납",
        "정리",
        "함",
    ) and has_any(text, "칸", "케이스", "보관", "정리", "보빈"):
        return "정리함", ""

    if has_any(
        text,
        "군번줄",
        "키링",
        "키고리",
        "스트랩",
        "체인",
        "o링",
        "oring",
        "메탈링",
        "악세",
        "팔찌",
        "목걸이",
        "핸드폰줄",
        "휴대폰",
    ):
        return "악세사리", ""

    if has_any(
        text,
        "e6000",
        "e-6000",
        "접착",
        "본드",
        "glue",
        "가와구찌",
        "올풀림",
    ) and not has_any(text, "b6000", "b7000"):
        return "기본부자재", "챠코페이퍼,접착제,기타제품"

    if has_any(
        text,
        "코바늘",
        "가위",
        "바늘",
        "골무",
        "차코",
        "chaco",
        "펠트",
        "직조",
        "확대경",
        "돋보기",
        "프리미",
        "와이어",
        "와이어",
        "테이프",
        "끈",
        "실끊",
        "실감개",
        "재봉",
        "단추",
        "단추",
        "단추",
        "핀",
        "옷핀",
        "구름솜",
        "솜",
        "마커",
        "펜",
        "자",
        "줄자",
        "핀셋",
        "집게",
        "와이어",
    ):
        return "기본부자재", basic_sub(text)

    if has_any(text, "십자수", "자수"):
        if has_any(text, "보석", "캔버스", "큐빅"):
            return "보석십자수 도안형", theme_subgroup(text)
        return "십자수패키지", theme_subgroup(text, kids_ok=False)

    return "기본부자재", "챠코페이퍼,접착제,기타제품"


def group_label(main: str, sub: str) -> str:
    if sub:
        return f"{main} - {sub}"
    return main


def load_products() -> pd.DataFrame:
    df = pd.read_excel(SRC)
    df = df[df["판매상태"].astype(str).str.contains("판매", na=False)]
    return df


def build_rows(df: pd.DataFrame) -> list[dict]:
    out = []
    for _, row in df.iterrows():
        name = clean(row.get("상품명", ""))
        opt = clean(row.get("옵션", "")) if "옵션" in df.columns else ""
        if not opt and "옵션명" in df.columns:
            opt = clean(row.get("옵션명", ""))
        cat = clean(row.get("카테고리", "")) or clean(row.get("사이트카테고리", ""))
        main, sub = classify(name, opt, cat)
        grp = group_label(main, sub)
        out.append(
            {
                "마스터상품번호": row.get("마스터상품번호", ""),
                "상품번호": row.get("상품번호", ""),
                "상품명": name,
                "판매가": parse_price(row.get("판매가", 0)),
                "판매상태": clean(row.get("판매상태", "판매중")) or "판매중",
                "그룹명칭": grp,
                "카테고리": cat,
                "대분류": main,
                "소분류": sub,
                "사이트": clean(row.get("사이트", "")),
            }
        )
    return out


def write_upload_xlsx(rows: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "그룹관리업로드"
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)
    ws.append(TEMPLATE_COLS)
    for c in ws[1]:
        c.fill = hf
        c.font = hfont
    for r in rows:
        ws.append(
            [
                r["마스터상품번호"],
                r["상품번호"],
                r["상품명"],
                f"{r['판매가']:.4f}" if r["판매가"] else "0.0000",
                r["판매상태"],
                r["그룹명칭"],
                r["카테고리"],
            ]
        )
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["F"].width = 42

    # 그룹별 상품 목록
    by_group: dict[str, list] = defaultdict(list)
    for r in rows:
        by_group[r["그룹명칭"]].append(r)

    ws2 = wb.create_sheet("그룹별목록")
    ws2.append(["그룹명칭", "건수", "마스터상품번호", "상품명", "사이트"])
    for grp in sorted(by_group.keys()):
        items = by_group[grp]
        for i, r in enumerate(sorted(items, key=lambda x: str(x["상품명"]))):
            ws2.append(
                [grp if i == 0 else "", len(items) if i == 0 else "", r["마스터상품번호"], r["상품명"], r["사이트"]]
            )

    # 마스터 기준 (중복 제거 — 지마켓/옥션 1행)
    seen = set()
    ws3 = wb.create_sheet("마스터기준")
    ws3.append(TEMPLATE_COLS + ["사이트"])
    for r in rows:
        mid = str(r["마스터상품번호"])
        if mid in seen:
            continue
        seen.add(mid)
        ws3.append(
            [
                r["마스터상품번호"],
                r["상품번호"],
                r["상품명"],
                f"{r['판매가']:.4f}" if r["판매가"] else "0.0000",
                r["판매상태"],
                r["그룹명칭"],
                r["카테고리"],
                r["사이트"],
            ]
        )
    wb.save(OUT_XLSX)


def write_summary(rows: list[dict]):
    cnt = Counter(r["그룹명칭"] for r in rows)
    master_cnt = Counter()
    seen_m = set()
    for r in rows:
        mid = str(r["마스터상품번호"])
        if mid in seen_m:
            continue
        seen_m.add(mid)
        master_cnt[r["그룹명칭"]] += 1

    lines = [
        f"지마켓/옥션 그룹관리 분류 ({TODAY})",
        f"원본: {SRC.name}",
        f"판매중 행: {len(rows)} / 마스터 기준: {len(seen_m)}개",
        "",
        "※ 그룹명칭 컬럼을 ESM 그룹관리 업로드 양식에 맞게 사용하세요.",
        "※ 지마켓·옥션 동일 마스터는 같은 그룹으로 지정됨.",
        "",
        "[그룹별 마스터 상품 수]",
        f"{'그룹명칭':<45} {'마스터':>6} {'행수':>6}",
        "-" * 60,
    ]
    for grp, n in master_cnt.most_common():
        lines.append(f"{grp:<45} {n:>6} {cnt[grp]:>6}")
    lines.extend(["", f"총 그룹 수: {len(master_cnt)}"])
    OUT_SUMMARY.write_text("\n".join(lines), encoding="utf-8")


def main():
    df = load_products()
    rows = build_rows(df)
    write_upload_xlsx(rows)
    write_summary(rows)
    print(f"Saved: {OUT_XLSX}")
    print(f"Saved: {OUT_SUMMARY}")
    print(f"Rows: {len(rows)}, masters: {len({str(r['마스터상품번호']) for r in rows})}, groups: {len({r['그룹명칭'] for r in rows})}")


if __name__ == "__main__":
    main()
