# -*- coding: utf-8 -*-
"""STIX MD - Chrome 기록 + 공개 상품페이지 기반 실데이터 수집"""
import json
import re
import shutil
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from playwright.sync_api import sync_playwright

OUT = Path(__file__).parent
TODAY = datetime.now().strftime("%Y-%m-%d")
HIST = Path.home() / "AppData/Local/Google/Chrome/User Data/Default/History"
TMP = OUT / "_hist_tmp"


def clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def read_history() -> list[tuple]:
    TMP.mkdir(exist_ok=True)
    dst = TMP / "History"
    shutil.copy2(HIST, dst)
    con = sqlite3.connect(dst)
    cur = con.cursor()
    cur.execute(
        """SELECT url, title, datetime(last_visit_time/1000000-11644473600,'unixetime','localtime')
        FROM urls ORDER BY last_visit_time DESC LIMIT 3000"""
    )
    # fix typo unixetime -> unixepoch
    cur.execute(
        """SELECT url, title, datetime(last_visit_time/1000000-11644473600,'unixepoch','localtime') as v
        FROM urls ORDER BY last_visit_time DESC LIMIT 3000"""
    )
    rows = cur.fetchall()
    con.close()
    return rows


def classify_urls(rows):
    out = {"coupang_products": [], "smartstore": [], "11st": [], "wing_edit": [], "esm": [], "cafe24": []}
    seen = set()
    for url, title, v in rows:
        if not url or url in seen:
            continue
        u = url.lower()
        if "coupang.com/vp/products/" in u:
            seen.add(url)
            out["coupang_products"].append({"url": url, "title": title, "date": v})
        elif "smartstore.naver.com" in u and "/products/" in u:
            seen.add(url)
            out["smartstore"].append({"url": url, "title": title, "date": v})
        elif "11st.co.kr/products/" in u:
            seen.add(url)
            out["11st"].append({"url": url, "title": title, "date": v})
        elif "wing.coupang.com" in u and ("vendor-inventory" in u or "modify" in u):
            seen.add(url)
            out["wing_edit"].append({"url": url, "title": title, "date": v})
        elif "esmplus.com/home/v2" in u:
            seen.add(url)
            out["esm"].append({"url": url, "title": title, "date": v})
        elif "escall.cafe24.com" in u:
            seen.add(url)
            out["cafe24"].append({"url": url, "title": title, "date": v})
    return out


def scrape_product_page(page, url: str) -> dict:
    rec = {"url": url, "title": "", "price": "", "seller": "", "ok": False}
    try:
        page.goto(url, timeout=60000, wait_until="domcontentloaded")
        page.wait_for_timeout(3000)
        if "coupang.com" in url:
            for sel in ["h1.product-title", "h1", ".prod-buy-header__title"]:
                if page.locator(sel).count():
                    rec["title"] = clean(page.locator(sel).first.inner_text(timeout=3000))
                    break
            if page.locator(".total-price strong").count():
                rec["price"] = clean(page.locator(".total-price strong").first.inner_text())
        elif "smartstore" in url:
            if page.locator("h3.DCVBehA8ZB").count():
                rec["title"] = clean(page.locator("h3.DCVBehA8ZB").first.inner_text())
            elif page.locator("h3").count():
                rec["title"] = clean(page.locator("h3").first.inner_text())
        elif "11st" in url:
            if page.locator("h1.title").count():
                rec["title"] = clean(page.locator("h1.title").first.inner_text())
            elif page.locator("h1").count():
                rec["title"] = clean(page.locator("h1").first.inner_text())
        rec["ok"] = bool(rec["title"])
    except Exception as e:
        rec["error"] = str(e)
    return rec


def seo_issues(title: str, platform: str) -> tuple[str, str]:
    problems, fixes = [], []
    t = title or ""
    if len(t) < 15:
        problems.append("상품명 너무 짧음")
        fixes.append("핵심키워드+사이즈+용도 추가")
    if platform == "coupang" and len(t) > 50:
        problems.append("쿠팡 50자 초과")
        fixes.append("핵심키워드만 남기고 축약")
    if platform == "gmarket" and len(t) < 40:
        problems.append("지마켓 키워드 부족")
        fixes.append("유사키워드·카테고리어 추가")
    if "보석십자수" not in t and "십자수" not in t:
        problems.append("핵심키워드 누락")
        fixes.append("보석십자수/DIY 키워드 앞쪽 배치")
    if not re.search(r"\d+x\d+", t, re.I):
        problems.append("사이즈 미표기")
        fixes.append("40x50 또는 30x40 표기")
    if platform == "smartstore" and t.count(",") > 2:
        problems.append("키워드 나열형")
        fixes.append("스팃스+자연문장형으로 변경")
    return (" / ".join(problems) if problems else "양호"), (" / ".join(fixes) if fixes else "-")


def write_report(urls, products, improvements):
    wb = Workbook()
    hf, hfnt = PatternFill("solid", fgColor="7A5C46"), Font(bold=True, color="FFFFFF")

    ws = wb.active
    ws.title = "Chrome기록_요약"
    ws.append(["구분", "건수", "비고"])
    for c in ws[1]:
        c.font, c.fill = hfnt, hf
    ws.append(["쿠팡 상품URL", len(urls["coupang_products"]), "주문/관리 이력"])
    ws.append(["스마트스토어", len(urls["smartstore"]), ""])
    ws.append(["11번가", len(urls["11st"]), ""])
    ws.append(["Wing 편집", len(urls["wing_edit"]), "vendorInventoryId 포함"])
    ws.append(["ESM", len(urls["esm"]), "escall 계정"])
    ws.append(["카페24 escall", len(urls["cafe24"]), "자사몰"])

    ws2 = wb.create_sheet("우리상품_스크랩")
    ws2.append(["플랫폼", "상품명", "가격", "URL", "SEO문제", "개선안"])
    for c in ws2[1]:
        c.font, c.fill = hfnt, hf
    for p in products:
        plat = "쿠팡" if "coupang" in p["url"] else "스마트스토어" if "smartstore" in p["url"] else "11번가"
        prob, fix = seo_issues(p.get("title", ""), plat.lower())
        ws2.append([plat, p.get("title", ""), p.get("price", ""), p["url"], prob, fix])

    ws3 = wb.create_sheet("Wing_편집이력")
    ws3.append(["날짜", "URL"])
    for c in ws3[1]:
        c.font, c.fill = hfnt, hf
    for w in urls["wing_edit"][:30]:
        ws3.append([w["date"], w["url"]])

    ws4 = wb.create_sheet("개선안_우선순위")
    ws4.append(["우선순위", "상품명", "문제점", "개선안", "플랫폼"])
    for c in ws4[1]:
        c.font, c.fill = hfnt, hf
    for row in improvements:
        ws4.append(row)

    for sh in wb.worksheets:
        for col in sh.columns:
            sh.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 2, 90)

    path = OUT / f"STIX_MD_실데이터_{TODAY}.xlsx"
    wb.save(path)
    (OUT / f"STIX_MD_실데이터_{TODAY}.json").write_text(
        json.dumps({"urls": urls, "products": products, "improvements": improvements}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def main():
    rows = read_history()
    urls = classify_urls(rows)

    # scrape unique product pages (max 12)
    targets = []
    for u in urls["coupang_products"][:6]:
        targets.append(u["url"].split("?")[0])
    for u in urls["smartstore"][:3]:
        targets.append(u["url"].split("?")[0])
    for u in urls["11st"][:3]:
        targets.append(u["url"].split("?")[0])
    targets = list(dict.fromkeys(targets))[:12]

    products = []
    with sync_playwright() as p:
        browser = p.chromium.launch(channel="chrome", headless=True)
        page = browser.new_page()
        for url in targets:
            print("scrape", url[:60])
            products.append(scrape_product_page(page, url))
            time.sleep(1)
        browser.close()

    improvements = []
    for i, p in enumerate(products):
        if not p.get("title"):
            continue
        plat = "coupang" if "coupang" in p["url"] else "smartstore" if "smartstore" in p["url"] else "gmarket"
        prob, fix = seo_issues(p["title"], plat)
        pri = "A" if prob != "양호" else "C"
        improvements.append([pri, p["title"], prob, fix, plat])

    path = write_report(urls, products, improvements)
    print("saved", path)
    print("products scraped:", sum(1 for p in products if p.get("ok")))
    return 0


if __name__ == "__main__":
    sys.exit(main())
