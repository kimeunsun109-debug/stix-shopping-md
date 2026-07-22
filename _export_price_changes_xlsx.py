# -*- coding: utf-8 -*-
"""Export Item Winner price-change history to Excel."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
HISTORY = ROOT / "item_winner" / "monitor_history.jsonl"

SKU_LABEL = {
    "p1_15mlx1": "1번 B7000 15ml×1 (vs 빙고)",
    "p2_15mlx3": "2번 B7000 15ml×3 (vs 온라인)",
    "p3_110mlx2": "3번 B7000 110ml×2 (vs 온라인)",
}
ACTION_KO = {"LOWER": "인하", "RAISE": "인상", "HOLD": "유지"}

thin = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Malgun Gothic", size=11)
cell_font = Font(name="Malgun Gothic", size=10)
title_font = Font(bold=True, name="Malgun Gothic", size=14, color="1F4E79")
sub_font = Font(name="Malgun Gothic", size=9, color="667085")
green = PatternFill("solid", fgColor="E8F5E9")
red = PatternFill("solid", fgColor="FFEBEE")
yellow = PatternFill("solid", fgColor="FFF8E1")
MONEY = "#,##0"


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin


def fmt_row(ws, r: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(r, c)
        cell.font = cell_font
        cell.border = thin
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> Path:
    rows: list[dict] = []
    if HISTORY.exists():
        for ln in HISTORY.read_text(encoding="utf-8").splitlines():
            if ln.strip():
                rows.append(json.loads(ln))

    prev: dict[str, int] = {}
    price_changes: list[dict] = []
    for r in rows:
        sku, my = r.get("sku"), r.get("my_price")
        if my is None or not sku:
            continue
        if sku in prev and prev[sku] != my:
            price_changes.append(
                {
                    "ts": r.get("ts", ""),
                    "sku": sku,
                    "prev": prev[sku],
                    "new": my,
                    "delta": my - prev[sku],
                    "comp": r.get("competitor_price"),
                    "action": r.get("action"),
                    "recommended": r.get("recommended"),
                    "reason": r.get("reason", ""),
                }
            )
        prev[sku] = my

    applies = [r for r in rows if r.get("applied")]
    latest: dict[str, dict] = {}
    for r in rows:
        sku = r.get("sku")
        if sku and r.get("my_price") is not None:
            latest[sku] = r

    first_price: dict[str, int] = {}
    change_cnt = {s: 0 for s in SKU_LABEL}
    for r in rows:
        sku, my = r.get("sku"), r.get("my_price")
        if sku in SKU_LABEL and my is not None and sku not in first_price:
            first_price[sku] = my
    for ch in price_changes:
        change_cnt[ch["sku"]] = change_cnt.get(ch["sku"], 0) + 1

    wb = Workbook()

    # --- 요약 ---
    ws0 = wb.active
    ws0.title = "요약"
    ws0["A1"] = "STIX B7000 아이템위너 가격변동 요약"
    ws0["A1"].font = title_font
    ws0["A2"] = f"작성시각 {datetime.now():%Y-%m-%d %H:%M}  |  출처 monitor_history.jsonl"
    ws0["A2"].font = sub_font

    for i, h in enumerate(
        ["상품", "시작 관측가", "현재 관측가", "누적증감", "변동횟수", "최근 경쟁가"], 1
    ):
        ws0.cell(4, i, h)
    style_header(ws0, 4, 6)

    r = 5
    for sku, label in SKU_LABEL.items():
        fp = first_price.get(sku)
        lp = (latest.get(sku) or {}).get("my_price")
        ws0.cell(r, 1, label)
        ws0.cell(r, 2, fp)
        ws0.cell(r, 3, lp)
        ws0.cell(r, 4, (lp - fp) if fp is not None and lp is not None else None)
        ws0.cell(r, 5, change_cnt.get(sku, 0))
        ws0.cell(r, 6, (latest.get(sku) or {}).get("competitor_price"))
        for c in (2, 3, 4, 6):
            ws0.cell(r, c).number_format = MONEY
        if ws0.cell(r, 4).value is not None and ws0.cell(r, 4).value < 0:
            ws0.cell(r, 4).fill = green
        elif ws0.cell(r, 4).value is not None and ws0.cell(r, 4).value > 0:
            ws0.cell(r, 4).fill = red
        fmt_row(ws0, r, 6)
        r += 1

    ws0["A9"] = "주요 실제 변동 타임라인"
    ws0["A9"].font = Font(bold=True, name="Malgun Gothic", size=12, color="1F4E79")
    ws0["A10"] = "일시"
    ws0["B10"] = "내용"
    style_header(ws0, 10, 2)
    timeline = [
        (
            "2026-07-16 17:07",
            "Wing 반영 시도: p1 1,650→1,500 / p2 10,560→10,290 / p3 13,800→13,790 "
            "(이후 p3는 13,800 유지 정책으로 복구)",
        ),
        ("2026-07-16 18:12", "관측: p2 10,560→10,290 반영 확인"),
        ("2026-07-20 12:06", "관측: p2 10,290→10,190 (시장/경쟁 변동)"),
        (
            "2026-07-20 12:24~12:45",
            "Wing 로그인 후 반영: p1 1,500 / p2 10,090 / p3 13,690",
        ),
        (
            "2026-07-20 12:45",
            "관측 확인: p1 1,650→1,500 / p2 10,190→10,090 / p3 13,800→13,690",
        ),
    ]
    tr = 11
    for t, desc in timeline:
        ws0.cell(tr, 1, t)
        ws0.cell(tr, 2, desc)
        fmt_row(ws0, tr, 2)
        tr += 1

    ws0["A17"] = "시트 안내"
    ws0["A17"].font = Font(bold=True, name="Malgun Gothic", size=12)
    guides = [
        "1) 요약 — 상품별 시작가/현재가",
        "2) 실제가격변동 — 모니터가 관측한 내 판매가 변경 (핵심)",
        "3) 현재가스냅샷 — SKU별 최신 1건",
        "4) Wing반영성공로그 — applied=True 기록 (중복 가능)",
        "5) 판단이력_인하인상 — LOWER/RAISE 판단 전체",
    ]
    for i, g in enumerate(guides):
        ws0.cell(18 + i, 1, g).font = cell_font
    autosize(ws0, [22, 14, 14, 12, 10, 14])
    ws0.column_dimensions["B"].width = 78

    # --- 실제가격변동 ---
    ws1 = wb.create_sheet("실제가격변동")
    ws1["A1"] = "STIX B7000 — 실제 관측 가격변동"
    ws1["A1"].font = title_font
    ws1["A2"] = "내 판매가가 직전 관측과 달라진 시점만 표시"
    ws1["A2"].font = sub_font
    headers1 = ["일시", "상품", "이전가", "변경후", "증감", "경쟁가", "판단", "권장가", "비고"]
    for i, h in enumerate(headers1, 1):
        ws1.cell(4, i, h)
    style_header(ws1, 4, len(headers1))
    r = 5
    for ch in price_changes:
        ws1.cell(r, 1, ch["ts"].replace("T", " "))
        ws1.cell(r, 2, SKU_LABEL.get(ch["sku"], ch["sku"]))
        ws1.cell(r, 3, ch["prev"])
        ws1.cell(r, 4, ch["new"])
        ws1.cell(r, 5, ch["delta"])
        ws1.cell(r, 6, ch["comp"])
        ws1.cell(r, 7, ACTION_KO.get(ch["action"], ch["action"] or ""))
        ws1.cell(r, 8, ch["recommended"])
        if ch["delta"] < 0:
            ws1.cell(r, 5).fill = green
            ws1.cell(r, 9, "인하 반영(관측)")
        elif ch["delta"] > 0:
            ws1.cell(r, 5).fill = red
            ws1.cell(r, 9, "인상 반영(관측)")
        for c in (3, 4, 5, 6, 8):
            ws1.cell(r, c).number_format = MONEY
        fmt_row(ws1, r, len(headers1))
        r += 1
    if not price_changes:
        ws1.cell(5, 1, "관측된 가격 변동 없음")
    autosize(ws1, [20, 36, 12, 12, 10, 12, 10, 12, 18])
    ws1.freeze_panes = "A5"

    # --- 현재가 ---
    ws2 = wb.create_sheet("현재가스냅샷")
    ws2["A1"] = "SKU별 최신 관측 스냅샷"
    ws2["A1"].font = title_font
    headers2 = ["상품", "내 판매가", "경쟁가", "위너", "판단", "권장가", "최종관측", "사유"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(3, i, h)
    style_header(ws2, 3, len(headers2))
    r = 4
    for sku in SKU_LABEL:
        x = latest.get(sku)
        if not x:
            continue
        ws2.cell(r, 1, SKU_LABEL[sku])
        ws2.cell(r, 2, x.get("my_price"))
        ws2.cell(r, 3, x.get("competitor_price"))
        ws2.cell(r, 4, "Y" if x.get("is_winner") else "N")
        ws2.cell(r, 5, ACTION_KO.get(x.get("action"), x.get("action") or ""))
        ws2.cell(r, 6, x.get("recommended"))
        ws2.cell(r, 7, (x.get("ts") or "").replace("T", " "))
        ws2.cell(r, 8, x.get("reason") or "")
        for c in (2, 3, 6):
            ws2.cell(r, c).number_format = MONEY
        fmt_row(ws2, r, len(headers2))
        r += 1
    autosize(ws2, [36, 12, 12, 8, 10, 12, 20, 48])
    ws2.freeze_panes = "A4"

    # --- Wing 반영 ---
    ws3 = wb.create_sheet("Wing반영성공로그")
    ws3["A1"] = "모니터 applied=True 기록"
    ws3["A1"].font = title_font
    ws3["A2"] = "일부 구간은 UI 미검증 중복 로그가 있을 수 있음. 실제 변동은 '실제가격변동' 시트 우선."
    ws3["A2"].font = sub_font
    headers3 = ["일시", "상품", "관측가(당시)", "경쟁가", "권장가", "반영메시지"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(4, i, h)
    style_header(ws3, 4, len(headers3))
    r = 5
    for x in applies:
        ws3.cell(r, 1, (x.get("ts") or "").replace("T", " "))
        ws3.cell(r, 2, SKU_LABEL.get(x.get("sku"), x.get("sku")))
        ws3.cell(r, 3, x.get("my_price"))
        ws3.cell(r, 4, x.get("competitor_price"))
        ws3.cell(r, 5, x.get("recommended"))
        msg = x.get("apply_msg") or ""
        if len(msg) > 90:
            msg = msg[:90] + "..."
        ws3.cell(r, 6, msg)
        for c in (3, 4, 5):
            ws3.cell(r, c).number_format = MONEY
        fmt_row(ws3, r, len(headers3))
        r += 1
    if not applies:
        ws3.cell(5, 1, "applied=True 기록 없음")
    autosize(ws3, [20, 36, 14, 12, 12, 55])
    ws3.freeze_panes = "A5"

    # --- 판단 이력 ---
    ws4 = wb.create_sheet("판단이력_인하인상")
    ws4["A1"] = "모니터 판단 LOWER/RAISE (유지 제외)"
    ws4["A1"].font = title_font
    headers4 = ["일시", "상품", "내가격", "경쟁가", "판단", "권장가", "반영", "반영결과", "사유"]
    for i, h in enumerate(headers4, 1):
        ws4.cell(3, i, h)
    style_header(ws4, 3, len(headers4))
    r = 4
    for x in rows:
        if x.get("action") not in ("LOWER", "RAISE"):
            continue
        ws4.cell(r, 1, (x.get("ts") or "").replace("T", " "))
        ws4.cell(r, 2, SKU_LABEL.get(x.get("sku"), x.get("sku")))
        ws4.cell(r, 3, x.get("my_price"))
        ws4.cell(r, 4, x.get("competitor_price"))
        ws4.cell(r, 5, ACTION_KO.get(x.get("action"), x.get("action")))
        ws4.cell(r, 6, x.get("recommended"))
        ws4.cell(r, 7, "Y" if x.get("applied") else "N")
        msg = x.get("apply_msg") or ""
        if len(msg) > 60:
            msg = msg[:60] + "..."
        ws4.cell(r, 8, msg)
        ws4.cell(r, 9, x.get("reason") or "")
        ws4.cell(r, 5).fill = green if x.get("action") == "LOWER" else yellow
        for c in (3, 4, 6):
            ws4.cell(r, c).number_format = MONEY
        fmt_row(ws4, r, len(headers4))
        r += 1
    autosize(ws4, [20, 36, 12, 12, 10, 12, 8, 40, 45])
    ws4.freeze_panes = "A4"

    out = ROOT / f"MD_아이템위너_가격변동_{datetime.now():%Y-%m-%d}.xlsx"
    wb.save(out)
    print(out)
    print(
        f"price_changes={len(price_changes)} applies={len(applies)} "
        f"decisions={sum(1 for x in rows if x.get('action') in ('LOWER','RAISE'))}"
    )
    return out


if __name__ == "__main__":
    main()
