# -*- coding: utf-8 -*-
"""STIX MD - 플랫폼별 상위노출 키워드/상품 스캔 (공개 검색 기준)"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import sync_playwright

OUT_DIR = Path(__file__).parent
TODAY = datetime.now().strftime("%Y-%m-%d")

KEYWORDS = [
    "보석십자수",
    "보석십자수 키트",
    "DIY 보석십자수",
    "어린이 보석십자수",
    "십자수 패키지",
    "수예용품 DIY",
]

OUR_BRANDS = ("스팃스", "STIX", "stix", "스팟스")


def clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def is_ours(title: str) -> bool:
    t = title.lower()
    return any(b.lower() in t for b in OUR_BRANDS)


def scan_coupang(page, keyword: str, max_items: int = 15) -> list[dict]:
    url = f"https://www.coupang.com/np/search?q={keyword}"
    page.goto(url, wait_until="networkidle", timeout=90000)
    page.wait_for_timeout(4000)
    # 차단/캡차 감지
    body = page.content()
    if "Access Denied" in body or "captcha" in body.lower() or "로봇" in body:
        page.screenshot(path=str(OUT_DIR / f"debug_coupang_{keyword[:8]}.png"))
        return []
    items = []
    selectors = [
        "li.search-product",
        "li[class*='search-product']",
        "ul#productList li",
        "[data-product-id]",
    ]
    cards = []
    for sel in selectors:
        cards = page.locator(sel).all()
        if cards:
            break
    for i, card in enumerate(cards[:max_items]):
        try:
            name_el = card.locator("div.name, .name").first
            price_el = card.locator("strong.price-value, .price-value").first
            rank = i + 1
            title = clean(name_el.inner_text(timeout=2000)) if name_el.count() else ""
            price = clean(price_el.inner_text(timeout=2000)) if price_el.count() else ""
            if not title:
                continue
            items.append(
                {
                    "platform": "쿠팡",
                    "keyword": keyword,
                    "rank": rank,
                    "title": title,
                    "price": price,
                    "is_ours": is_ours(title),
                }
            )
        except Exception:
            continue
    return items


def scan_naver(page, keyword: str, max_items: int = 15) -> list[dict]:
    url = f"https://search.shopping.naver.com/search/all?query={keyword}"
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(3000)
    items = []
    cards = page.locator("[class*='product_item'], .product_item__MDtDF, div[class*='product_info']").all()
    if not cards:
        cards = page.locator("div[class*='basicList_item']").all()
    for i, card in enumerate(cards[:max_items]):
        try:
            title_el = card.locator("a[class*='product_link'], .product_title, a[title]").first
            price_el = card.locator("[class*='price'], .price_num").first
            title = ""
            if title_el.count():
                title = clean(title_el.get_attribute("title") or title_el.inner_text(timeout=2000))
            price = clean(price_el.inner_text(timeout=2000)) if price_el.count() else ""
            if not title:
                continue
            items.append(
                {
                    "platform": "네이버쇼핑",
                    "keyword": keyword,
                    "rank": i + 1,
                    "title": title,
                    "price": price,
                    "is_ours": is_ours(title),
                }
            )
        except Exception:
            continue
    return items


def extract_title_patterns(rows: list[dict]) -> dict:
    """상위 노출 상품명에서 반복 키워드 빈도"""
    words: dict[str, int] = {}
    for r in rows:
        for w in re.split(r"[\s,/|+\[\]()]+", r["title"]):
            w = w.strip()
            if len(w) >= 2 and not w.isdigit():
                words[w] = words.get(w, 0) + 1
    top = sorted(words.items(), key=lambda x: -x[1])[:30]
    return dict(top)


def write_excel(all_rows: list[dict], patterns: dict, our_rows: list[dict]):
    wb = Workbook()

    # Sheet 1: 검색결과 원본
    ws1 = wb.active
    ws1.title = "검색결과_원본"
    headers = ["플랫폼", "검색키워드", "순위", "상품명", "가격", "우리상품여부"]
    ws1.append(headers)
    for h in range(1, len(headers) + 1):
        c = ws1.cell(1, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="7A5C46")
        c.alignment = Alignment(horizontal="center")
    for r in all_rows:
        ws1.append([r["platform"], r["keyword"], r["rank"], r["title"], r["price"], "Y" if r["is_ours"] else ""])

    # Sheet 2: 우리 상품 노출
    ws2 = wb.create_sheet("우리상품_노출")
    ws2.append(["플랫폼", "검색키워드", "순위", "상품명", "가격", "비고"])
    for h in range(1, 7):
        ws2.cell(1, h).font = Font(bold=True)
    if our_rows:
        for r in our_rows:
            note = "상위권" if r["rank"] <= 10 else "중하위"
            ws2.append([r["platform"], r["keyword"], r["rank"], r["title"], r["price"], note])
    else:
        ws2.append(["-", "-", "-", "스캔 키워드 내 스팃스/STIX 노출 없음", "-", "상품명·브랜드 확인 필요"])

    # Sheet 3: 키워드 패턴
    ws3 = wb.create_sheet("상위노출_키워드")
    ws3.append(["키워드", "빈도", "활용제안"])
    for h in range(1, 4):
        ws3.cell(1, h).font = Font(bold=True)
    suggestions = {
        "보석십자수": "핵심키워드 - 모든 플랫폼 상품명 앞 15자 내",
        "DIY": "쿠팡·11번가 필수",
        "캔버스형": "형태 구분 - 옵션명에도 반영",
        "액자형": "지마켓·옥션 검색어 강화",
        "키트": "세트/구성품 강조",
        "40x50": "사이즈 숫자형 노출",
        "30x40": "입문용 사이즈 키워드",
        "초보자": "전환 키워드",
        "취미": "집콕·힐링과 병행",
        "어린이": "키즈 라인 분리",
        "스티커": "어린이 보석십자수 서브키워드",
        "해바라기": "풍수·인테리어 트렌드",
        "디즈니": "캐릭터 IP - 매칭 검토",
        "명화": "고가 라인 키워드",
    }
    for kw, cnt in patterns.items():
        ws3.append([kw, cnt, suggestions.get(kw, "")])

    # Sheet 4: 작업대기 - 상품명 개선
    ws4 = wb.create_sheet("상품명_개선대기")
    ws4.append(["우선순위", "플랫폼", "현재상품명", "문제점", "개선안", "승인"])
    for h in range(1, 7):
        ws4.cell(1, h).font = Font(bold=True)
    ws4.append(["A", "쿠팡", "(Wing 카탈로그 연동 후 입력)", "브랜드명만 있거나 키워드 부족 가능", "핵심키워드+용도+사이즈+타겟", ""])
    ws4.append(["A", "지마켓", "(ESM 카탈로그 연동 후 입력)", "검색어 10개 미만", "유사키워드·카테고리·소재 추가", ""])
    ws4.append(["B", "스마트스토어", "(스마트스토어 연동 후 입력)", "키워드 나열형", "스팃스+자연문장형 상품명", ""])

    # Sheet 5: 썸네일/상세 개선
    ws5 = wb.create_sheet("썸네일_상세_개선")
    ws5.append(["우선순위", "상품명", "유형", "문제점", "개선구조", "승인"])
    for h in range(1, 7):
        ws5.cell(1, h).font = Font(bold=True)
    ws5.append([
        "A", "보석십자수 공통", "썸네일",
        "완성작만 있고 구성품/사이즈 미표시",
        "1장:완성작+사이즈배지 / 2장:구성품 / 3장:작업장면 / 4장:인테리어",
        "",
    ])
    ws5.append([
        "A", "보석십자수 공통", "상세페이지",
        "첫화면 매력도·FAQ 부족",
        "4분할 HTML(히어로→구성→과정→FAQ) + alt텍스트 SEO",
        "",
    ])

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    out = OUT_DIR / f"MD_시장분석_{TODAY}.xlsx"
    wb.save(out)
    return out


def main():
    all_rows: list[dict] = []
    with sync_playwright() as p:
        browser = p.chromium.launch(channel="chrome", headless=True)
        context = browser.new_context(
            locale="ko-KR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()
        for kw in KEYWORDS[:4]:
            print(f"[쿠팡] {kw}...", flush=True)
            try:
                rows = scan_coupang(page, kw)
                all_rows.extend(rows)
                print(f"  -> {len(rows)}건", flush=True)
            except Exception as e:
                print(f"  -> 실패: {e}", flush=True)
        for kw in KEYWORDS[:2]:
            print(f"[네이버] {kw}...", flush=True)
            try:
                rows = scan_naver(page, kw)
                all_rows.extend(rows)
                print(f"  -> {len(rows)}건", flush=True)
            except Exception as e:
                print(f"  -> 실패: {e}", flush=True)
        browser.close()

    our_rows = [r for r in all_rows if r["is_ours"]]
    patterns = extract_title_patterns(all_rows)
    out_xlsx = write_excel(all_rows, patterns, our_rows)

    json_out = OUT_DIR / f"MD_시장분석_{TODAY}.json"
    json_out.write_text(json.dumps({"date": TODAY, "rows": all_rows, "patterns": patterns}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"저장: {out_xlsx}")
    print(f"저장: {json_out}")
    print(f"총 {len(all_rows)}건, 우리상품 {len(our_rows)}건")
    return 0 if all_rows else 1


if __name__ == "__main__":
    sys.exit(main())
