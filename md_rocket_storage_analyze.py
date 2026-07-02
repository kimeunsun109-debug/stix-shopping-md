# -*- coding: utf-8 -*-
"""로켓그로스 보관비 분석 - 회수 vs 가격조정 판매 vs 유지관망"""
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent
STORAGE_XLSX = BASE / "로켓그로스 보관비.xlsx"
COUPANG_XLSX = BASE / "쇼핑몰별 전체상품" / "쿠팡.xlsx"
TODAY = datetime.now().strftime("%Y-%m-%d")
OUT = BASE / f"STIX_로켓그로스_보관비_분류_{TODAY}.xlsx"
TXT = BASE / f"MD_로켓그로스_보관비_분류_{TODAY}.txt"

FREE_DAYS = 30


def clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def num(v, default=0.0):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def estimate_price(name: str, opt: str) -> int:
    text = f"{name} {opt}".upper()
    if "B7000" in text or "E6000" in text or "B6000" in text:
        if "110" in text and ("× 4" in opt or "X 4" in text or "4개" in opt):
            return 42000
        if "110" in text:
            return 18000
        if "15ML" in text and ("× 5" in opt or "5개" in opt):
            return 12000
        if "15ML" in text and ("× 3" in opt or "3개" in opt):
            return 9000
        return 15000
    if "스티커" in name or "어린이" in name:
        if "액자" in name:
            return 15900
        return 9900
    if "프린트" in name and "십자수" in name:
        return 18900
    if "해바라기" in name:
        if "40×80" in opt or "50×90" in opt or "40×70" in opt:
            return 32000
        return 27900
    if "재물" in name or "부엉이" in opt or "백호" in opt:
        return 25900
    if "도구" in name and "세트" in name:
        return 22900
    if "보석십자수" in name:
        if "40×50" in opt or "40X50" in opt:
            return 29900
        return 24900
    return 20000


def read_storage_rows():
    wb = openpyxl.load_workbook(STORAGE_XLSX, read_only=True, data_only=True)
    rows = list(wb["보관비"].iter_rows(min_row=9, values_only=True))
    wb.close()
    return rows


def read_coupang_status():
    wb = openpyxl.load_workbook(COUPANG_XLSX, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [clean(h) for h in all_rows[3]]
    idx = {h: i for i, h in enumerate(headers)}
    out = {}
    for row in all_rows[4:]:
        if not any(row):
            continue
        oid = clean(row[idx["옵션 ID"]]) if "옵션 ID" in idx else ""
        if not oid:
            continue
        out[oid] = {
            "status": clean(row[idx.get("판매상태", -1)] if "판매상태" in idx else ""),
            "pid": clean(row[idx.get("등록상품ID", -1)] if "등록상품ID" in idx else ""),
        }
    return out


def aggregate_storage(rows):
    by_sku = defaultdict(
        lambda: {
            "fee_week": 0.0,
            "max_days": 0,
            "max_qty": 0,
            "name": "",
            "opt": "",
            "pid": "",
            "oid": "",
            "cbm": 0.0,
            "saver": False,
            "display_date": "",
        }
    )
    for r in rows:
        if not r or not r[9]:
            continue
        sku = str(r[9])
        rec = by_sku[sku]
        rec["fee_week"] += num(r[19])
        rec["max_days"] = max(rec["max_days"], int(num(r[6])))
        rec["max_qty"] = max(rec["max_qty"], int(num(r[12])))
        rec["name"] = clean(r[10])
        rec["opt"] = clean(r[11])
        rec["pid"] = clean(r[7])
        rec["oid"] = clean(r[8])
        rec["cbm"] = max(rec["cbm"], num(r[13]))
        if r[14] == "Y":
            rec["saver"] = True
        rec["display_date"] = clean(r[5]) or rec["display_date"]
    return by_sku


def is_core_sku(name: str, opt: str) -> bool:
    t = f"{name} {opt}".upper()
    return any(
        k in t
        for k in (
            "B7000", "E6000", "B6000", "접착제",
            "어린이 보석십자수 스티커", "STIX 어린이",
        )
    )


def is_slow_sku(name: str, opt: str) -> bool:
    t = f"{name} {opt}"
    slow = ("해바라기", "프린트 실 십자수", "재물을 불러오는", "프랑스자수", "펠트", "원목")
    return any(k in t for k in slow)


def classify(rec, status: str):
    price = estimate_price(rec["name"], rec["opt"])
    stock_val = rec["max_qty"] * price
    fee = rec["fee_week"]
    days = rec["max_days"]
    fee_month = fee * 4.3
    fee_per_unit_week = fee / rec["max_qty"] if rec["max_qty"] else fee
    fee_ratio_month = (fee_month / stock_val * 100) if stock_val else 0

    reasons = []
    action = ""
    suggest_price = ""
    discount_pct = 0

    # 무료 구간 — 아직 보관비 없음
    if fee <= 0:
        if days <= FREE_DAYS:
            action = "유지_관망"
            reasons.append(f"무료보관 구간({days}일)")
            if rec["max_qty"] >= 20:
                reasons.append(f"재고 {rec['max_qty']}개 — 30일 경과 전 판매 촉진 권장")
        else:
            action = "유지_관망"
            reasons.append(f"보관비 0원(세이버/할인) — {days}일")
        guide = "30일 경과 전 판매 모니터링 / 재고 과다 시 선제 할인"
        urgency = "낮음"
        return _pack(action, urgency, price, stock_val, fee_month, fee_ratio_month, suggest_price, reasons, guide, discount_pct)

    # --- 보관비 발생 ---
    urgency = "긴급" if fee >= 400 or days >= 120 else ("높음" if fee >= 100 or days >= 90 else "보통")

    # 1) 회수(빼기): 장기체화 + 저회전 + 소량 또는 보관비 부담 큼
    remove_score = 0
    if days >= 120:
        remove_score += 3
        reasons.append(f"장기체화 {days}일")
    elif days >= 90:
        remove_score += 2
        reasons.append(f"체화 {days}일")

    if is_slow_sku(rec["name"], rec["opt"]):
        remove_score += 2
        reasons.append("저회전 도안/카테고리")

    if rec["max_qty"] <= 5 and days >= 80:
        remove_score += 2
        reasons.append(f"소량재고 {rec['max_qty']}개")

    if fee_ratio_month >= 3:
        remove_score += 2
        reasons.append(f"월보관비/재고가치 {fee_ratio_month:.1f}%")

    if fee_per_unit_week >= 15 and not is_core_sku(rec["name"], rec["opt"]):
        remove_score += 1
        reasons.append(f"개당 주보관비 {fee_per_unit_week:.0f}원")

    # 2) 가격조정 판매: 핵심 SKU / 재고 많음 / 할인으로 회전 가능
    sell_score = 0
    if is_core_sku(rec["name"], rec["opt"]):
        sell_score += 3
        reasons.append("핵심/회전 SKU")
    if rec["max_qty"] >= 10:
        sell_score += 2
        reasons.append(f"재고 {rec['max_qty']}개")
    if stock_val >= 150000:
        sell_score += 1
        reasons.append(f"재고가치 약 {stock_val:,}원")
    if 60 <= days < 120 and fee > 0:
        sell_score += 1
        reasons.append("보관비 초기 — 빠른 처리 유리")

    # 해바라기·프린트십자수·재물시리즈 소량 옵션 → 회수 우선
    if is_slow_sku(rec["name"], rec["opt"]) and rec["max_qty"] <= 10 and days >= 100:
        remove_score += 2

    # B7000 대용량 등 접착제 → 할인 판매
    if "B7000" in rec["name"].upper() and rec["max_qty"] >= 10:
        sell_score += 3

    # 어린이 스티커/액자 — 묶음 할인
    if "어린이" in rec["name"] and rec["max_qty"] >= 8:
        sell_score += 2

    if remove_score >= sell_score + 2:
        action = "회수_빼기"
        guide = "Wing > 재고회수 신청 (로켓그로스 출고). 타 채널 재고 있으면 거기서만 판매"
    elif sell_score >= remove_score:
        action = "가격조정_판매"
        if days >= 120 or fee_ratio_month >= 2.5:
            discount_pct = 25
        elif days >= 90:
            discount_pct = 20
        elif days >= 60:
            discount_pct = 15
        else:
            discount_pct = 10
        if is_core_sku(rec["name"], rec["opt"]) and "접착제" in rec["name"]:
            discount_pct = min(discount_pct, 15)
        suggest_price = max(1000, int(price * (100 - discount_pct) / 100 / 100) * 100)
        guide = (
            f"판매가 {price:,}원 → {suggest_price:,}원 ({discount_pct}%↓) / "
            "쿠팡 즉시할인·로켓와우쿠폰·검색광고 2주 테스트"
        )
    else:
        action = "회수_우선_할인검토"
        discount_pct = 20
        suggest_price = max(1000, int(price * 0.80 / 100) * 100)
        guide = f"2주간 {discount_pct}% 할인 테스트 → 미판매 시 회수 ({suggest_price:,}원)"

    return _pack(action, urgency, price, stock_val, fee_month, fee_ratio_month, suggest_price, reasons, guide, discount_pct)


def _pack(action, urgency, price, stock_val, fee_month, fee_ratio, suggest_price, reasons, guide, discount_pct):
    return {
        "action": action,
        "urgency": urgency,
        "price": price,
        "stock_val": stock_val,
        "est_month_fee": round(fee_month),
        "fee_ratio_pct": round(fee_ratio, 1),
        "suggest_price": suggest_price,
        "discount_pct": discount_pct,
        "reason": " / ".join(dict.fromkeys(reasons)),
        "guide": guide,
    }


def write_report(items, summary):
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    fills = {
        "회수_빼기": PatternFill("solid", fgColor="FCE4D6"),
        "회수_우선_할인검토": PatternFill("solid", fgColor="FFF2CC"),
        "가격조정_판매": PatternFill("solid", fgColor="E2EFDA"),
        "유지_관망": PatternFill("solid", fgColor="DDEBF7"),
    }
    action_label = {
        "회수_빼기": "① 회수(빼기)",
        "회수_우선_할인검토": "② 회수우선·할인검토",
        "가격조정_판매": "③ 가격조정·판매",
        "유지_관망": "④ 유지·관망",
    }

    cols = [
        "분류", "긴급도", "SKU ID", "옵션ID", "등록상품ID", "등록상품명", "옵션명",
        "보관재고", "보관기간(일)", "이번주보관비", "추정월보관비", "월보관비/재고%",
        "추정판매가", "재고가치(추정)", "권장판매가", "할인율", "판매상태", "세이버",
        "선정사유", "조치안",
    ]

    sheets = [
        ("1_회수_빼기", lambda x: x["action"] == "회수_빼기", lambda x: (-x["fee_week"], -x["max_days"])),
        ("2_가격조정_판매", lambda x: x["action"] == "가격조정_판매", lambda x: (-x["fee_week"], -x["stock_val"])),
        ("3_회수우선_할인검토", lambda x: x["action"] == "회수_우선_할인검토", lambda x: (-x["fee_week"],)),
        ("4_유지_관망", lambda x: x["action"] == "유지_관망", lambda x: (-x["max_qty"], -x["max_days"])),
        ("5_전체_SKU", lambda x: True, lambda x: (-x["fee_week"], -x["max_days"])),
    ]

    first = True
    for sheet_name, filt, sort_key in sheets:
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        ws.append(cols)
        for c in ws[1]:
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        for it in sorted([x for x in items if filt(x)], key=sort_key):
            ws.append([
                action_label.get(it["action"], it["action"]),
                it["urgency"], it["sku"], it["oid"], it["pid"], it["name"], it["opt"],
                it["max_qty"], it["max_days"], it["fee_week"], it["est_month_fee"], it["fee_ratio_pct"],
                it["price"], it["stock_val"], it["suggest_price"] or "",
                f"{it['discount_pct']}%" if it["discount_pct"] else "",
                it["status"], "Y" if it["saver"] else "N",
                it["reason"], it["guide"],
            ])
            fill = fills.get(it["action"])
            if fill and sheet_name != "5_전체_SKU":
                for c in ws[ws.max_row]:
                    c.fill = fill

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(42, max(10, len(str(col[0].value or "")) + 2))

    ws_sum = wb.create_sheet("0_요약", 0)
    ws_sum.append(["항목", "내용"])
    for k, v in summary.items():
        ws_sum.append([k, v])
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 55

    wb.save(OUT)


def write_txt(items, summary):
    lines = [f"STIX 로켓그로스 보관비 분류 ({TODAY})", "=" * 50, ""]
    for k, v in summary.items():
        lines.append(f"{k}: {v}")
    lines.append("")
    for action, title in [
        ("회수_빼기", "① 회수(빼기) — Wing 재고회수"),
        ("가격조정_판매", "③ 가격조정·판매 — 할인 후 2주 테스트"),
        ("회수_우선_할인검토", "② 회수우선·할인검토"),
        ("유지_관망", "④ 유지·관망 — 무료구간/모니터링"),
    ]:
        grp = [x for x in items if x["action"] == action]
        grp.sort(key=lambda x: -x.get("fee_week", 0))
        lines.append(f"\n[{title}] {len(grp)}건 / 이번주 보관비 {sum(x['fee_week'] for x in grp):,.0f}원")
        for x in grp[:20]:
            fee_s = f" 보관비 {x['fee_week']:.0f}원" if x["fee_week"] else ""
            lines.append(f"  · {x['name'][:35]} | {x['opt'][:30]}{fee_s}")
            if x.get("guide"):
                lines.append(f"    → {x['guide'][:80]}")
        if len(grp) > 20:
            lines.append(f"  ... 외 {len(grp)-20}건 (엑셀 참조)")
    TXT.write_text("\n".join(lines), encoding="utf-8")


def main():
    rows = read_storage_rows()
    by_sku = aggregate_storage(rows)
    status_map = read_coupang_status()

    items = []
    total_fee = 0.0
    for sku, rec in by_sku.items():
        st = status_map.get(rec["oid"], {})
        cls = classify(rec, st.get("status", ""))
        total_fee += rec["fee_week"]
        items.append({
            "sku": sku,
            "oid": rec["oid"],
            "pid": rec["pid"],
            "name": rec["name"],
            "opt": rec["opt"],
            "max_qty": rec["max_qty"],
            "max_days": rec["max_days"],
            "fee_week": rec["fee_week"],
            "saver": rec["saver"],
            "status": st.get("status", ""),
            **cls,
        })

    counts = {a: len([x for x in items if x["action"] == a]) for a in (
        "회수_빼기", "가격조정_판매", "회수_우선_할인검토", "유지_관망"
    )}
    fee_by_action = {a: sum(x["fee_week"] for x in items if x["action"] == a) for a in counts}

    summary = {
        "분석일": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "원본": STORAGE_XLSX.name,
        "정산주기": "2026-06-28 (주간)",
        "이번주 보관비 합계": f"{total_fee:,.0f}원 (VAT별도, 파일 기준)",
        "전체 SKU": len(items),
        "① 회수(빼기)": f"{counts['회수_빼기']}건 / 보관비 {fee_by_action['회수_빼기']:,.0f}원",
        "③ 가격조정·판매": f"{counts['가격조정_판매']}건 / 보관비 {fee_by_action['가격조정_판매']:,.0f}원",
        "② 회수우선·할인검토": f"{counts['회수_우선_할인검토']}건 / 보관비 {fee_by_action['회수_우선_할인검토']:,.0f}원",
        "④ 유지·관망": f"{counts['유지_관망']}건 (무료구간·보관비 0)",
        "판매가": "쿠팡 export에 가격 없음 → 카테고리별 추정가 사용",
        "무료보관": f"진열 후 약 {FREE_DAYS}일 (데이터 기준)",
    }

    write_report(items, summary)
    write_txt(items, summary)
    print(f"Saved: {OUT}")
    print(f"Saved: {TXT}")
    for k, v in counts.items():
        print(k, v)


if __name__ == "__main__":
    main()
